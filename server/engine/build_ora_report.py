from __future__ import annotations

import json
import math
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from calendar import monthrange
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


BASE = Path(os.environ.get("ORA_BASE", "/input"))
WORK = Path(os.environ.get("ORA_WORK", "/work"))
OUT_DIR = WORK / "outputs"

TEMPLATE = BASE / "ora外送周报模版.xlsx"
PREVIOUS = BASE / "ORA外送周报_6.8-6.14.xlsx"

START = pd.Timestamp(os.environ.get("ORA_START", "2026-06-15"))
END = pd.Timestamp(os.environ.get("ORA_END", "2026-06-21"))
PREV_START = pd.Timestamp(os.environ.get("ORA_PREV_START", START - pd.Timedelta(days=7)))
PREV_END = pd.Timestamp(os.environ.get("ORA_PREV_END", END - pd.Timedelta(days=7)))


def sheet_label(start: pd.Timestamp, end: pd.Timestamp) -> str:
    return f"{start.month}.{start.day}-{end.month}.{end.day}"


def export_label(start: pd.Timestamp, end: pd.Timestamp) -> str:
    is_full_month_range = (
        start.day == 1
        and end.day == monthrange(end.year, end.month)[1]
        and end >= start
    )
    if is_full_month_range:
        if start.year == end.year and start.month == end.month:
            return f"{start.month}月"
        if start.year == end.year:
            return f"{start.month}-{end.month}月"
        return f"{start.year}.{start.month}-{end.year}.{end.month}月"
    return sheet_label(start, end)


CURRENT_SHEET = os.environ.get("ORA_CURRENT_SHEET", sheet_label(START, END))
PREVIOUS_SHEET = os.environ.get("ORA_PREVIOUS_SHEET", sheet_label(PREV_START, PREV_END))
PERIOD_DAYS = int((END - START).days) + 1
TEMPLATE_CURRENT_LABELS = ("6.15-6.21",)
TEMPLATE_PREVIOUS_LABELS = ("6.8-6.14",)


@dataclass
class Store:
    name_full: str
    ele_id: str
    mt_id: str
    name: str
    code: str


_EXCEL_CACHE: dict[tuple[str, tuple[tuple[str, str], ...]], Any] = {}


def _excel_cache_key(name: str, kwargs: dict[str, Any]) -> tuple[str, tuple[tuple[str, str], ...]]:
    return name, tuple(sorted((key, repr(value)) for key, value in kwargs.items()))


def _copy_excel_data(value: Any) -> Any:
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, dict):
        return {key: df.copy() if isinstance(df, pd.DataFrame) else df for key, df in value.items()}
    return value


def read_excel(name: str, **kwargs) -> pd.DataFrame:
    key = _excel_cache_key(name, kwargs)
    if key not in _EXCEL_CACHE:
        _EXCEL_CACHE[key] = pd.read_excel(BASE / name, **kwargs)
    return _copy_excel_data(_EXCEL_CACHE[key])


def read_excel_columns(name: str, columns: dict[str, list[str]], sheet_name: str | int = 0, header_rows: int = 30) -> pd.DataFrame:
    """Read only selected columns from an xlsx sheet by matching header text."""
    key = _excel_cache_key(
        f"columns:{name}",
        {
            "sheet_name": sheet_name,
            "columns": tuple((target, tuple(aliases)) for target, aliases in columns.items()),
        },
    )
    if key in _EXCEL_CACHE:
        return _copy_excel_data(_EXCEL_CACHE[key])

    wb = load_workbook(BASE / name, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[int(sheet_name)]
        header_row: int | None = None
        selected: dict[str, int] = {}
        normalized_aliases = {
            target: [norm_header(alias) for alias in aliases]
            for target, aliases in columns.items()
        }
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=header_rows, values_only=True), start=1):
            header_map = {
                norm_header(value): idx
                for idx, value in enumerate(row)
                if norm_header(value)
            }
            candidate: dict[str, int] = {}
            for target, aliases in normalized_aliases.items():
                for alias in aliases:
                    if alias in header_map:
                        candidate[target] = header_map[alias]
                        break
            if len(candidate) == len(columns):
                header_row = row_idx
                selected = candidate
                break
        if header_row is None:
            missing = ", ".join(columns)
            raise KeyError(f"{name} 缺少必要表头：{missing}")

        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            record: dict[str, Any] = {}
            non_empty = False
            for target, idx in selected.items():
                value = row[idx] if idx < len(row) else None
                record[target] = value
                if value not in (None, ""):
                    non_empty = True
            if non_empty:
                rows.append(record)
    finally:
        wb.close()

    result = pd.DataFrame(rows, columns=list(columns))
    _EXCEL_CACHE[key] = result
    return result.copy()


def norm_id(value: Any) -> str:
    if pd.isna(value):
        return ""
    try:
        return str(int(float(value)))
    except Exception:
        return str(value).strip()


def norm_store_name(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"ora\s*coffee", "", text, flags=re.IGNORECASE)
    text = text.replace("咖啡", "").replace("店", "")
    return re.sub(r"[\s·•\.\-_/\\()（）【】\[\]{}]+", "", text)


def to_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.replace("--", 0), errors="coerce").fillna(0)


def parse_date_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_datetime(series.astype("Int64").astype(str), format="%Y%m%d", errors="coerce")
    return pd.to_datetime(series, errors="coerce")


def period_rows(df: pd.DataFrame, date_col: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    out = df.copy()
    out["_date"] = parse_date_series(out[date_col])
    return out[(out["_date"] >= start) & (out["_date"] <= end)].copy()


def current_rows(df: pd.DataFrame, date_col: str) -> pd.DataFrame:
    return period_rows(df, date_col, START, END)


def safe_div(num: Any, den: Any) -> float | None:
    try:
        n = float(num)
        d = float(den)
    except Exception:
        return None
    if not math.isfinite(n) or not math.isfinite(d) or abs(d) < 1e-12:
        return None
    return n / d


def scalar_num(value: Any) -> float:
    try:
        result = float(pd.to_numeric(value, errors="coerce"))
    except Exception:
        return 0.0
    return result if math.isfinite(result) else 0.0


def growth(cur: Any, prev: Any) -> float | None:
    div = safe_div(cur, prev)
    return None if div is None else div - 1


def diff(cur: Any, prev: Any) -> float | None:
    try:
        if cur is None or prev is None:
            return None
        return float(cur) - float(prev)
    except Exception:
        return None


def pct(value: Any) -> float:
    if pd.isna(value) or value in ("", "-"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if abs(float(value)) <= 1 else float(value) / 100
    text = str(value).strip().replace("+", "")
    if text in ("", "-"):
        return 0.0
    if text.endswith("%"):
        return float(text[:-1]) / 100
    try:
        number = float(text)
        return number if abs(number) <= 1 else number / 100
    except Exception:
        return 0.0


def write(cell, value: Any) -> None:
    if isinstance(value, float) and not math.isfinite(value):
        value = None
    cell.value = value


def clean_error_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("#"):
        return None
    return value


def get_prev(ws, row: int, col: int) -> Any:
    return ws.cell(row, col).value


SECTION_TITLES = ("营业数据", "流量数据", "推广数据", "门店评分")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sheet_store_row_maps(sheet, start: int, end: int, code_col: int, name_col: int) -> tuple[dict[str, int], dict[str, int]]:
    by_code: dict[str, int] = {}
    by_name: dict[str, int] = {}
    for row in range(start, end + 1):
        code = str(sheet.cell(row, code_col).value or "").strip()
        name = norm_store_name(sheet.cell(row, name_col).value)
        if code:
            by_code.setdefault(code, row)
        if name:
            by_name.setdefault(name, row)
    return by_code, by_name


def find_section_title_row(sheet, title: str) -> int | None:
    max_col = min(sheet.max_column or 1, 8)
    for row in range(1, (sheet.max_row or 1) + 1):
        for col in range(1, max_col + 1):
            if cell_text(sheet.cell(row, col).value) == title:
                return row
    return None


def find_next_section_title_row(sheet, section_row: int) -> int | None:
    max_col = min(sheet.max_column or 1, 8)
    for row in range(section_row + 1, (sheet.max_row or 1) + 1):
        for col in range(1, max_col + 1):
            if cell_text(sheet.cell(row, col).value) in SECTION_TITLES:
                return row
    return None


def find_section_header_row(sheet, title: str, fallback: int, code_col: int = 2, name_col: int = 3) -> int:
    section_row = find_section_title_row(sheet, title)
    if not section_row:
        return max(fallback - 1, 1)
    next_section = find_next_section_title_row(sheet, section_row) or ((sheet.max_row or section_row) + 1)
    for row in range(section_row + 1, next_section):
        if cell_text(sheet.cell(row, code_col).value) == "店号" and cell_text(sheet.cell(row, name_col).value) == "店名":
            return row
    return max(fallback - 1, section_row + 1)


def find_section_total_row(sheet, title: str, fallback: int) -> int:
    section_row = find_section_title_row(sheet, title)
    if not section_row:
        return fallback
    next_section = find_next_section_title_row(sheet, section_row) or ((sheet.max_row or section_row) + 1)
    for row in range(section_row + 1, next_section):
        for col in range(1, 5):
            if cell_text(sheet.cell(row, col).value) == "总计":
                return row
    return fallback


def section_store_row_maps(
    sheet,
    title: str,
    fallback_start: int,
    fallback_end: int,
    code_col: int = 2,
    name_col: int = 3,
) -> tuple[dict[str, int], dict[str, int]]:
    section_row = find_section_title_row(sheet, title)
    if not section_row:
        return sheet_store_row_maps(sheet, fallback_start, fallback_end, code_col, name_col)
    header_row = find_section_header_row(sheet, title, fallback_start, code_col, name_col)
    total_row = find_section_total_row(sheet, title, fallback_end + 1)
    next_section = find_next_section_title_row(sheet, section_row) or ((sheet.max_row or section_row) + 1)
    start = header_row + 1
    end = min(total_row - 1, next_section - 1)
    if end < start:
        return sheet_store_row_maps(sheet, fallback_start, fallback_end, code_col, name_col)
    return sheet_store_row_maps(sheet, start, end, code_col, name_col)


def matched_row_by_store(store: Store, by_code: dict[str, int], by_name: dict[str, int]) -> int | None:
    return (
        by_code.get(store.code)
        or by_name.get(norm_store_name(store.name))
        or by_name.get(norm_store_name(store.name_full))
    )


def matched_row_by_code_name(code: Any, name: Any, by_code: dict[str, int], by_name: dict[str, int]) -> int | None:
    code_key = str(code or "").strip()
    name_key = norm_store_name(name)
    return by_code.get(code_key) or by_name.get(name_key)


def store_code_by_sheet_row(sheet, row: int, stores: list[Store], code_col: int, name_col: int) -> str:
    known_codes = {store.code for store in stores}
    code = str(sheet.cell(row, code_col).value or "").strip()
    if code in known_codes:
        return code
    name = norm_store_name(sheet.cell(row, name_col).value)
    by_name: dict[str, str] = {}
    for store in stores:
        by_name[norm_store_name(store.name)] = store.code
        by_name[norm_store_name(store.name_full)] = store.code
    return by_name.get(name, code)


def delivery_row_maps(sheet, start: int = 3, end: int = 18) -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    by_mt_id: dict[str, int] = {}
    by_ele_id: dict[str, int] = {}
    by_name: dict[str, int] = {}
    for row in range(start, end + 1):
        mt_id = norm_id(sheet.cell(row, 4).value)
        ele_id = norm_id(sheet.cell(row, 3).value)
        name = norm_store_name(sheet.cell(row, 2).value)
        if mt_id:
            by_mt_id.setdefault(mt_id, row)
        if ele_id:
            by_ele_id.setdefault(ele_id, row)
        if name:
            by_name.setdefault(name, row)
    return by_mt_id, by_ele_id, by_name


def matched_delivery_row(sheet, row: int, maps: tuple[dict[str, int], dict[str, int], dict[str, int]]) -> int | None:
    by_mt_id, by_ele_id, by_name = maps
    mt_id = norm_id(sheet.cell(row, 4).value)
    ele_id = norm_id(sheet.cell(row, 3).value)
    name = norm_store_name(sheet.cell(row, 2).value)
    return by_mt_id.get(mt_id) or by_ele_id.get(ele_id) or by_name.get(name)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def build_store_list() -> tuple[list[Store], dict[str, str], dict[str, str]]:
    df = read_excel("ORA门店信息表.xlsx").dropna(how="all")
    stores: list[Store] = []
    mt_to_code: dict[str, str] = {}
    ele_to_code: dict[str, str] = {}
    for _, row in df.iterrows():
        if pd.isna(row.get("店号")):
            continue
        store = Store(
            name_full=str(row["门店名称"]).strip(),
            ele_id=norm_id(row["饿了么门店ID"]),
            mt_id=norm_id(row["美团门店ID"]),
            name=str(row["店名"]).strip(),
            code=str(row["店号"]).strip(),
        )
        stores.append(store)
        mt_to_code[store.mt_id] = store.code
        ele_to_code[store.ele_id] = store.code
    return stores, mt_to_code, ele_to_code


def sum_by_store(df: pd.DataFrame, store_col: str, cols: list[str]) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for code, group in df.groupby(store_col):
        result[str(code)] = {col: float(to_num(group[col]).sum()) if col in group else 0.0 for col in cols}
    return result


def ensure_current_sheet(wb) -> None:
    if CURRENT_SHEET in wb.sheetnames:
        return
    preferred = ["当周页", "当周", "本期", "6.15-6.21", "6.8-6.14"]
    for name in preferred:
        if name in wb.sheetnames:
            wb[name].title = CURRENT_SHEET
            return
    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if len(visible) > 1:
        visible[1].title = CURRENT_SHEET


def previous_period_sheet(prev_wb):
    if PREVIOUS_SHEET in prev_wb.sheetnames:
        return prev_wb[PREVIOUS_SHEET]
    excluded = {"V2", "上期"}
    visible = [
        ws
        for ws in prev_wb.worksheets
        if getattr(ws, "sheet_state", "visible") == "visible" and ws.title not in excluded
    ]
    for ws in visible:
        if re.search(r"\d{1,2}\.\d{1,2}\s*[-~]\s*\d{1,2}\.\d{1,2}", ws.title):
            return ws
    if visible:
        return visible[0]
    return prev_wb[prev_wb.sheetnames[0]]


def sheet_layout_from_xlsx(path: Path, sheet_title: str) -> dict[str, Any]:
    ns_main = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    with zipfile.ZipFile(path) as zf:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("r:Relationship", ns_rel)
        }
        sheet_path = None
        for sheet in workbook.findall("m:sheets/m:sheet", ns_main):
            if sheet.attrib.get("name") == sheet_title:
                rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                target = rel_targets.get(rel_id or "")
                if target:
                    target = target.lstrip("/")
                    if target.startswith("xl/"):
                        sheet_path = target
                    elif target.startswith("../"):
                        sheet_path = target[3:]
                    else:
                        sheet_path = "xl/" + target
                break
        if not sheet_path:
            return {"cols": [], "rows": {}, "merges": [], "freeze": None}
        root = ET.fromstring(zf.read(sheet_path))

    cols = []
    for col in root.findall("m:cols/m:col", ns_main):
        cols.append(dict(col.attrib))

    rows = {}
    for row in root.findall("m:sheetData/m:row", ns_main):
        idx = row.attrib.get("r")
        if idx:
            rows[int(idx)] = dict(row.attrib)

    merges = [
        merge.attrib.get("ref")
        for merge in root.findall("m:mergeCells/m:mergeCell", ns_main)
        if merge.attrib.get("ref")
    ]

    pane = root.find("m:sheetViews/m:sheetView/m:pane", ns_main)
    freeze = None
    if pane is not None and pane.attrib.get("state") == "frozen":
        freeze = pane.attrib.get("topLeftCell")
    return {"cols": cols, "rows": rows, "merges": merges, "freeze": freeze}


def refresh_period_labels(wb) -> None:
    """Update visible template period captions to the selected comparison periods."""
    range_pattern = re.compile(r"\d{1,2}[./]\d{1,2}\s*[-~]\s*\d{1,2}[./]\d{1,2}")

    def replace_text(text: str, sheet_title: str) -> str:
        updated = text
        if "本期" in updated:
            return range_pattern.sub(CURRENT_SHEET, updated)
        if "上期" in updated:
            return range_pattern.sub(PREVIOUS_SHEET, updated)
        if sheet_title == "上期":
            return updated
        for label in TEMPLATE_CURRENT_LABELS:
            updated = updated.replace(label, CURRENT_SHEET)
        for label in TEMPLATE_PREVIOUS_LABELS:
            updated = updated.replace(label, PREVIOUS_SHEET)
        return updated

    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and not value.startswith("="):
                    refreshed = replace_text(value, ws.title)
                    if refreshed != value:
                        cell.value = refreshed


def apply_total_row_bold(wb) -> None:
    """Keep total rows visually distinct without changing other formatting."""
    fixed_rows = {
        "V2": [19, 39, 59, 80],
        CURRENT_SHEET: [19, 39, 59],
        "用户体验-客诉": [17],
    }

    def bold_row(ws, row: int) -> None:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            font = copy(cell.font)
            font.bold = True
            cell.font = font

    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        if ws.title == "上期":
            continue
        rows_to_bold = set(fixed_rows.get(ws.title, []))
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() in {"总计", "合计"}:
                    rows_to_bold.add(cell.row)
        for row in rows_to_bold:
            if 1 <= row <= ws.max_row:
                bold_row(ws, row)


def compute_metrics(stores: list[Store], mt_to_code: dict[str, str], ele_to_code: dict[str, str]) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    mt_store = current_rows(read_excel("美团门店数据.xlsx"), "日期")
    mt_store["_id"] = mt_store["门店id"].map(norm_id)
    mt_store["code"] = mt_store["_id"].map(mt_to_code)
    mt_store = mt_store[mt_store["code"].notna()].copy()

    ele_store = current_rows(read_excel("饿了么门店数据.xlsx"), "日期")
    ele_store["_id"] = ele_store["门店编号"].map(norm_id)
    ele_store["code"] = ele_store["_id"].map(ele_to_code)
    ele_store = ele_store[ele_store["code"].notna()].copy()

    days_orders: dict[tuple[str, pd.Timestamp], float] = defaultdict(float)
    for _, row in mt_store.iterrows():
        days_orders[(row["code"], row["_date"].normalize())] += scalar_num(row.get("有效订单", 0))
    for _, row in ele_store.iterrows():
        days_orders[(row["code"], row["_date"].normalize())] += scalar_num(row.get("有效订单", 0))
    biz_days = {s.code: 0 for s in stores}
    for (code, _dt), orders in days_orders.items():
        if orders > 0:
            biz_days[code] = biz_days.get(code, 0) + 1
    total_biz_days = sum(biz_days.values())

    mt_store["_sales"] = to_num(mt_store["营业收入"]) + to_num(mt_store["平台服务费(含佣金和配送服务费)"])
    ele_store["_sales"] = to_num(ele_store["收入"]) + to_num(ele_store["平台技术服务费"]) + to_num(ele_store["履约技术服务费"])
    mt_paid_exp_from_store = "曝光提升数(次)" in mt_store.columns
    ele_paid_exp_from_store = "曝光提升数" in ele_store.columns

    mt_ag = sum_by_store(
        mt_store,
        "code",
        ["_sales", "商家活动支出", "有效订单", "曝光人数", "入店人数", "下单人数", "曝光次数", "曝光提升数(次)", "优惠前总额"],
    )
    ele_ag = sum_by_store(
        ele_store,
        "code",
        ["_sales", "商家活动成本（含满减活动）", "有效订单", "曝光人数", "进店人数", "下单人数", "曝光次数", "曝光提升数", "营业额"],
    )

    mt_score = {}
    for code, group in mt_store.sort_values("_date").groupby("code"):
        mt_score[code] = float(to_num(group.tail(1)["综合体验分"]).iloc[0])
    ele_score = {}
    for code, group in ele_store.sort_values("_date").groupby("code"):
        ele_score[code] = float(to_num(group.tail(1)["店铺评分"]).iloc[0])

    mt_promo = current_rows(read_excel("美团推广.xlsx", sheet_name="效果数据"), "日期")
    mt_promo["_id"] = mt_promo["门店ID"].map(norm_id)
    mt_promo["code"] = mt_promo["_id"].map(mt_to_code)
    mt_exclude = ["津贴联盟", "赏金联盟", "流量助手", "金字招牌", "袋鼠店长", "品牌装修", "应用市场", "短信通", "拼好饭"]
    text_cols = [c for c in ["推广产品", "计划名称", "营销场景"] if c in mt_promo.columns]
    if text_cols:
        mask = pd.Series(False, index=mt_promo.index)
        for col in text_cols:
            mask |= mt_promo[col].astype(str).apply(lambda x: any(term in x for term in mt_exclude))
        mt_promo = mt_promo[~mask].copy()
    mt_promo = mt_promo[mt_promo["code"].notna()].copy()
    mt_pr = sum_by_store(
        mt_promo,
        "code",
        ["推广消费实付(元)", "曝光提升数(次)", "访问提升数(次)", "订单原价交易额(元)"],
    )

    ele_promo = current_rows(read_excel("饿了么推广.xlsx"), "日期")
    ele_promo["_id"] = ele_promo["门店ID"].map(norm_id)
    ele_promo["code"] = ele_promo["_id"].map(ele_to_code)
    text_cols = [c for c in ["推广产品", "计划名称"] if c in ele_promo.columns]
    if text_cols:
        mask = pd.Series(False, index=ele_promo.index)
        for col in text_cols:
            mask |= ele_promo[col].astype(str).str.contains("增量助手", na=False)
        ele_promo = ele_promo[~mask].copy()
    ele_promo = ele_promo[ele_promo["code"].notna()].copy()
    ele_pr = sum_by_store(
        ele_promo,
        "code",
        ["推广现金消费(元)", "曝光提升数", "进店提升数", "订单原价交易额(元)"],
    )

    praise = read_excel("好评数中差评数据.xlsx")
    praise["_id"] = praise["平台门店ID"].map(norm_id)
    praise["code_mt"] = praise["_id"].map(mt_to_code)
    praise["code_ele"] = praise["_id"].map(ele_to_code)
    mt_reviews = praise[praise["外卖平台"].astype(str).str.contains("美团", na=False)].copy()
    ele_reviews = praise[praise["外卖平台"].astype(str).str.contains("饿了么", na=False)].copy()
    mt_r = {
        str(code): {"good": float(to_num(g["好评数"]).sum()), "bad": float(to_num(g["中差评数"]).sum())}
        for code, g in mt_reviews[mt_reviews["code_mt"].notna()].groupby("code_mt")
    }
    ele_r = {
        str(code): {"good": float(to_num(g["好评数"]).sum()), "bad": float(to_num(g["中差评数"]).sum())}
        for code, g in ele_reviews[ele_reviews["code_ele"].notna()].groupby("code_ele")
    }

    metrics: dict[str, dict[str, Any]] = {}
    for store in stores:
        code = store.code
        mt = mt_ag.get(code, {})
        el = ele_ag.get(code, {})
        mpr = mt_pr.get(code, {})
        epr = ele_pr.get(code, {})
        bd = biz_days.get(code, 0)
        den = bd or None

        mt_sales = mt.get("_sales", 0.0)
        ele_sales = el.get("_sales", 0.0)
        mt_discount = mt.get("商家活动支出", 0.0)
        ele_discount = el.get("商家活动成本（含满减活动）", 0.0)
        mt_orders = mt.get("有效订单", 0.0)
        ele_orders = el.get("有效订单", 0.0)
        mt_exp = mt.get("曝光人数", 0.0)
        ele_exp = el.get("曝光人数", 0.0)
        mt_entry = mt.get("入店人数", 0.0)
        ele_entry = el.get("进店人数", 0.0)
        mt_buyers = mt.get("下单人数", 0.0)
        ele_buyers = el.get("下单人数", 0.0)
        mt_exp_count = mt.get("曝光次数", 0.0)
        ele_exp_count = el.get("曝光次数", 0.0)
        mt_base_gmv = mt.get("优惠前总额", 0.0)
        ele_base_gmv = el.get("营业额", 0.0)
        total_discount_base = mt_base_gmv + ele_base_gmv

        mt_paid_exp = mt.get("曝光提升数(次)", 0.0) if mt_paid_exp_from_store else mpr.get("曝光提升数(次)", 0.0)
        ele_paid_exp = el.get("曝光提升数", 0.0) if ele_paid_exp_from_store else epr.get("曝光提升数", 0.0)
        mt_spend = mpr.get("推广消费实付(元)", 0.0)
        ele_spend = epr.get("推广现金消费(元)", 0.0)
        mt_visit_lift = mpr.get("访问提升数(次)", 0.0)
        ele_visit_lift = epr.get("进店提升数", 0.0)
        mt_orig = mpr.get("订单原价交易额(元)", 0.0)
        ele_orig = epr.get("订单原价交易额(元)", 0.0)

        total_sales = mt_sales + ele_sales
        total_discount = mt_discount + ele_discount
        total_orders = mt_orders + ele_orders
        total_exp = mt_exp + ele_exp
        total_entry = mt_entry + ele_entry
        total_buyers = mt_buyers + ele_buyers
        total_exp_count = mt_exp_count + ele_exp_count
        total_paid_exp = mt_paid_exp + ele_paid_exp
        total_spend = mt_spend + ele_spend
        total_visit_lift = mt_visit_lift + ele_visit_lift
        total_orig = mt_orig + ele_orig

        score_vals = [v for v in [mt_score.get(code), ele_score.get(code)] if v not in (None, 0)]
        metrics[code] = {
            "store": store,
            "biz_days": bd,
            "total": {
                "sales": total_sales,
                "sales_daily": safe_div(total_sales, den),
                "discount": total_discount,
                "discount_base": total_discount_base,
                "discount_rate": safe_div(total_discount, total_discount_base),
                "orders": total_orders,
                "orders_daily": safe_div(total_orders, den),
                "at": safe_div(total_sales, total_orders),
                "exp_people": total_exp,
                "exp_people_daily": safe_div(total_exp, den),
                "entry_rate": safe_div(total_entry, total_exp),
                "order_rate": safe_div(total_buyers, total_entry),
                "exp_count": total_exp_count,
                "paid_exp": total_paid_exp,
                "ad_share": safe_div(total_paid_exp, total_exp_count),
                "ad_spend": total_spend,
                "ad_visits": total_visit_lift,
                "ad_roi": safe_div(total_orig, total_spend),
                "ad_orig": total_orig,
                "ad_gmv_share": safe_div(total_orig, mt_base_gmv + ele_base_gmv),
                "score": sum(score_vals) / len(score_vals) if score_vals else None,
                "good": mt_r.get(code, {}).get("good", 0.0) + ele_r.get(code, {}).get("good", 0.0),
                "bad": mt_r.get(code, {}).get("bad", 0.0) + ele_r.get(code, {}).get("bad", 0.0),
            },
            "mt": {
                "sales": mt_sales,
                "sales_daily": safe_div(mt_sales, den),
                "discount": mt_discount,
                "discount_base": mt_base_gmv,
                "discount_rate": safe_div(mt_discount, mt_base_gmv),
                "orders": mt_orders,
                "orders_daily": safe_div(mt_orders, den),
                "at": safe_div(mt_sales, mt_orders),
                "exp_people": mt_exp,
                "exp_people_daily": safe_div(mt_exp, den),
                "entry_rate": safe_div(mt_entry, mt_exp),
                "order_rate": safe_div(mt_buyers, mt_entry),
                "exp_count": mt_exp_count,
                "paid_exp": mt_paid_exp,
                "ad_share": safe_div(mt_paid_exp, mt_exp_count),
                "ad_spend": mt_spend,
                "ad_visits": mt_visit_lift,
                "ad_roi": safe_div(mt_orig, mt_spend),
                "ad_orig": mt_orig,
                "ad_gmv_share": safe_div(mt_orig, mt_base_gmv),
                "score": mt_score.get(code),
                "good": mt_r.get(code, {}).get("good", 0.0),
                "bad": mt_r.get(code, {}).get("bad", 0.0),
            },
            "ele": {
                "sales": ele_sales,
                "sales_daily": safe_div(ele_sales, den),
                "discount": ele_discount,
                "discount_base": ele_base_gmv,
                "discount_rate": safe_div(ele_discount, ele_base_gmv),
                "orders": ele_orders,
                "orders_daily": safe_div(ele_orders, den),
                "at": safe_div(ele_sales, ele_orders),
                "exp_people": ele_exp,
                "exp_people_daily": safe_div(ele_exp, den),
                "entry_rate": safe_div(ele_entry, ele_exp),
                "order_rate": safe_div(ele_buyers, ele_entry),
                "exp_count": ele_exp_count,
                "paid_exp": ele_paid_exp,
                "ad_share": safe_div(ele_paid_exp, ele_exp_count),
                "ad_spend": ele_spend,
                "ad_visits": ele_visit_lift,
                "ad_roi": safe_div(ele_orig, ele_spend),
                "ad_orig": ele_orig,
                "ad_gmv_share": safe_div(ele_orig, ele_base_gmv),
                "score": ele_score.get(code),
                "good": ele_r.get(code, {}).get("good", 0.0),
                "bad": ele_r.get(code, {}).get("bad", 0.0),
            },
        }

    totals = {
        "biz_days": total_biz_days,
        "mt_promo_codes": set(mt_promo["code"].unique()),
        "ele_promo_codes": set(ele_promo["code"].unique()),
        "mt_store_codes": set(mt_store["code"].unique()),
        "ele_store_codes": set(ele_store["code"].unique()),
        "paid_exp_source": {
            "美团": "美团门店数据.曝光提升数(次)" if mt_paid_exp_from_store else "美团推广.曝光提升数(次)（美团门店数据缺少该列）",
            "饿了么": "饿了么门店数据.曝光提升数" if ele_paid_exp_from_store else "饿了么推广.曝光提升数（饿了么门店数据缺少该列）",
        },
    }
    for scope in ["total", "mt", "ele"]:
        subtotal: dict[str, float] = defaultdict(float)
        for code in [s.code for s in stores]:
            for key, value in metrics[code][scope].items():
                if key in {"sales_daily", "discount_rate", "orders_daily", "at", "exp_people_daily", "entry_rate", "order_rate", "ad_share", "ad_roi", "ad_gmv_share", "score"}:
                    continue
                if isinstance(value, (int, float)):
                    subtotal[key] += float(value)
        den = total_biz_days or None
        subtotal["sales_daily"] = safe_div(subtotal["sales"], den)
        subtotal["discount_rate"] = safe_div(subtotal["discount"], subtotal.get("discount_base", 0.0))
        subtotal["orders_daily"] = safe_div(subtotal["orders"], den)
        subtotal["at"] = safe_div(subtotal["sales"], subtotal["orders"])
        subtotal["exp_people_daily"] = safe_div(subtotal["exp_people"], den)
        subtotal["entry_rate"] = safe_div(
            sum(metrics[c][scope].get("exp_people", 0) * (metrics[c][scope].get("entry_rate") or 0) for c in [s.code for s in stores]),
            subtotal["exp_people"],
        )
        # Use actual summed numerators for conversion rates.
        if scope == "mt":
            entry_sum = sum(mt_ag.get(s.code, {}).get("入店人数", 0.0) for s in stores)
            buyer_sum = sum(mt_ag.get(s.code, {}).get("下单人数", 0.0) for s in stores)
        elif scope == "ele":
            entry_sum = sum(ele_ag.get(s.code, {}).get("进店人数", 0.0) for s in stores)
            buyer_sum = sum(ele_ag.get(s.code, {}).get("下单人数", 0.0) for s in stores)
        else:
            entry_sum = sum(mt_ag.get(s.code, {}).get("入店人数", 0.0) + ele_ag.get(s.code, {}).get("进店人数", 0.0) for s in stores)
            buyer_sum = sum(mt_ag.get(s.code, {}).get("下单人数", 0.0) + ele_ag.get(s.code, {}).get("下单人数", 0.0) for s in stores)
        subtotal["entry_rate"] = safe_div(entry_sum, subtotal["exp_people"])
        subtotal["order_rate"] = safe_div(buyer_sum, entry_sum)
        subtotal["ad_share"] = safe_div(subtotal["paid_exp"], subtotal["exp_count"])
        subtotal["ad_roi"] = safe_div(subtotal["ad_orig"], subtotal["ad_spend"])
        if scope == "mt":
            base_gmv = sum(mt_ag.get(s.code, {}).get("优惠前总额", 0.0) for s in stores)
        elif scope == "ele":
            base_gmv = sum(ele_ag.get(s.code, {}).get("营业额", 0.0) for s in stores)
        else:
            base_gmv = sum(mt_ag.get(s.code, {}).get("优惠前总额", 0.0) + ele_ag.get(s.code, {}).get("营业额", 0.0) for s in stores)
        subtotal["ad_gmv_share"] = safe_div(subtotal["ad_orig"], base_gmv)
        totals[scope] = subtotal
    return metrics, totals


def compute_distance(mt_to_code: dict[str, str], ele_to_code: dict[str, str]) -> dict[str, dict[str, dict[str, float]]]:
    df = read_excel("ora_订单距离分布_2026-06-22.xlsx", sheet_name="订单分布")
    df["_id"] = df["平台门店ID"].map(norm_id)
    result: dict[str, dict[str, dict[str, float]]] = defaultdict(lambda: {"mt": {}, "ele": {}, "total": {}})
    for _, row in df.iterrows():
        platform = str(row["平台"])
        code = mt_to_code.get(row["_id"]) if platform == "美团" else ele_to_code.get(row["_id"])
        if not code:
            continue
        vals = {
            "0_1": pct(row.get("[0,0.5Km)")) + pct(row.get("[0.5,0.8Km)")) + pct(row.get("[0.8,1.0Km)")),
            "1_3": pct(row.get("[1.0,2.0Km)")) + pct(row.get("[2.0,3.0Km)")),
            "gt3": pct(row.get("[3.0,4.0Km)")) + pct(row.get("[4.0,5.0Km)")) + pct(row.get(">5.0Km")),
        }
        result[code]["mt" if platform == "美团" else "ele"] = vals
    for code, data in result.items():
        for key in ["0_1", "1_3", "gt3"]:
            parts = [data[scope].get(key) for scope in ("mt", "ele") if key in data[scope]]
            data["total"][key] = sum(parts) / len(parts) if parts else 0.0
    return result


def compute_paid_intervals(mt_to_code: dict[str, str], ele_to_code: dict[str, str], stores: list[Store] | None = None) -> tuple[dict[str, dict[str, dict[str, float]]], dict[str, Any]]:
    result: dict[str, dict[str, dict[str, float]]] = defaultdict(lambda: {"mt": {}, "ele": {}, "total": {}})
    audit: dict[str, Any] = {}
    if stores is None:
        stores = build_store_list()[0]

    code_lookup = {store.code.upper(): store.code for store in stores}
    store_name_items: list[tuple[str, str]] = []
    for store in stores:
        for name in [store.name, store.name_full, store.code]:
            key = norm_store_name(name)
            if key:
                store_name_items.append((key, store.code))

    def match_by_name(value: Any) -> str | None:
        key = norm_store_name(value)
        if not key:
            return None
        exact = {name_key: code for name_key, code in store_name_items}
        if key in exact:
            return exact[key]
        matches = [code for name_key, code in store_name_items if name_key and (name_key in key or key in name_key)]
        return matches[0] if len(set(matches)) == 1 else None

    def assign_codes(df: pd.DataFrame, platform: str) -> pd.Series:
        platform_map = mt_to_code if platform == "mt" else ele_to_code
        code = pd.Series([None] * len(df), index=df.index, dtype="object")

        id_cols = (
            ["美团门店ID", "门店id", "门店ID", "平台门店ID", "门店编号"]
            if platform == "mt"
            else ["饿了么门店ID", "门店编号", "门店id", "门店ID", "平台门店ID"]
        )
        match_counts: dict[str, int] = {}
        for col in id_cols:
            if col not in df.columns:
                continue
            mapped = df[col].map(norm_id).map(platform_map)
            fill_mask = code.isna() & mapped.notna()
            code.loc[fill_mask] = mapped.loc[fill_mask]
            match_counts[f"id:{col}"] = int(fill_mask.sum())

        for col in ["店号", "门店编码", "门店代码"]:
            if col not in df.columns:
                continue
            mapped = df[col].astype(str).str.strip().str.upper().map(code_lookup)
            fill_mask = code.isna() & mapped.notna()
            code.loc[fill_mask] = mapped.loc[fill_mask]
            match_counts[f"code:{col}"] = int(fill_mask.sum())

        for col in ["店名", "门店名称", "店铺名称", "外卖通门店名称"]:
            if col not in df.columns:
                continue
            mapped = df[col].apply(match_by_name)
            fill_mask = code.isna() & mapped.notna()
            code.loc[fill_mask] = mapped.loc[fill_mask]
            match_counts[f"name:{col}"] = int(fill_mask.sum())

        audit[f"{platform}_match_counts"] = match_counts
        return code

    def prepare_orders(fname: str, platform: str) -> pd.DataFrame:
        df = current_rows(read_excel(fname), "日期")
        df["code"] = assign_codes(df, platform)
        if "订单状态" in df.columns:
            if platform == "mt":
                df = df[df["订单状态"].isin(["已完成", "已接单"])].copy()
            else:
                df = df[df["订单状态"].eq("订单完结")].copy()
        paid_col = "订单实付" if platform == "mt" else "顾客实付"
        audit[f"{platform}_paid_column"] = paid_col if paid_col in df.columns else f"missing:{paid_col}"
        if paid_col not in df.columns:
            return df.iloc[0:0].copy()
        df["_paid"] = to_num(df[paid_col])
        return df[df["code"].notna()].copy()

    mt = prepare_orders("美团订单数据.xlsx", "mt")
    ele = prepare_orders("饿了么订单数据.xlsx", "ele")
    audit["mt_order_matched_rows"] = int(len(mt))
    audit["ele_order_matched_rows"] = int(len(ele))

    def counts(df: pd.DataFrame) -> dict[str, dict[str, float]]:
        out: dict[str, dict[str, float]] = {}
        for code, group in df.groupby("code"):
            paid = group["_paid"]
            c0 = int(((paid >= 0) & (paid < 20)).sum())
            c1 = int(((paid >= 20) & (paid < 30)).sum())
            c2 = int((paid >= 30).sum())
            total = c0 + c1 + c2
            out[str(code)] = {
                "0_20_count": c0,
                "20_30_count": c1,
                "gt30_count": c2,
                "total_count": total,
                "0_20": safe_div(c0, total) or 0.0,
                "20_30": safe_div(c1, total) or 0.0,
                "gt30": safe_div(c2, total) or 0.0,
            }
        return out

    mt_counts = counts(mt)
    ele_counts = counts(ele)
    all_codes = set(mt_counts) | set(ele_counts)
    for code in all_codes:
        result[code]["mt"] = mt_counts.get(code, {"0_20": 0.0, "20_30": 0.0, "gt30": 0.0, "total_count": 0, "0_20_count": 0, "20_30_count": 0, "gt30_count": 0})
        result[code]["ele"] = ele_counts.get(code, {"0_20": 0.0, "20_30": 0.0, "gt30": 0.0, "total_count": 0, "0_20_count": 0, "20_30_count": 0, "gt30_count": 0})
        total_count = result[code]["mt"]["total_count"] + result[code]["ele"]["total_count"]
        for bucket, count_key in [("0_20", "0_20_count"), ("20_30", "20_30_count"), ("gt30", "gt30_count")]:
            count = result[code]["mt"].get(count_key, 0) + result[code]["ele"].get(count_key, 0)
            result[code]["total"][bucket] = safe_div(count, total_count) or 0.0
    return result, audit


PACKAGE_CANONICAL = [
    "恰巴塔x拿铁套餐",
    "深刻拿铁（双杯）",
    "深刻拿铁+超大杯耶加雪菲美式",
    "恰巴塔x美式套餐",
    "超大杯美式·耶加雪菲（双杯）",
    "超大杯美式·红宝石瑰夏（双杯）",
    "车厘子可可拿铁（双杯）",
]
PACKAGE_ALIASES = {"超大杯美式·红宝石瑰夏（双杯套餐）": "超大杯美式·红宝石瑰夏（双杯）"}


def norm_product(name: str) -> str:
    return re.sub(r"\s+", "", str(name).strip().replace("(", "（").replace(")", "）"))


def canonical_package(name: Any) -> str | None:
    if pd.isna(name):
        return None
    raw = str(name).strip()
    n = norm_product(raw)
    if all(part in n for part in ["超大杯美式", "红宝石瑰夏", "双杯"]):
        return "超大杯美式·红宝石瑰夏（双杯）"
    for alias, canon in PACKAGE_ALIASES.items():
        if norm_product(alias) == n or norm_product(alias) in n:
            return canon
    for canon in PACKAGE_CANONICAL:
        cn = norm_product(canon)
        if cn == n or cn in n:
            return canon
    if "套餐" in raw or "双杯" in raw or "+" in raw or "＋" in raw:
        return raw
    return None


def infer_category(name: str, category_map: dict[str, str]) -> str:
    if name in category_map:
        return category_map[name]
    n = norm_product(name)
    for key, value in category_map.items():
        if norm_product(key) == n:
            return value
    if any(word in name for word in ["恰巴塔", "羊角", "贝果", "三明治"]):
        return "面包烘焙"
    if any(word in name for word in ["咖啡豆", "挂耳", "牛奶"]):
        return "零售"
    return "饮品"


def compute_products(prev_wb, total_store_days: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, float], dict[str, float], list[str]]:
    category_map: dict[str, str] = {}
    prev_single = prev_wb["商品销售排行-单品"]
    prev_single_total_row = find_label_row(prev_single, "总计", None, 2) or 63
    for row in range(2, prev_single_total_row):
        name = prev_single.cell(row, 3).value
        cat = prev_single.cell(row, 2).value
        if name and cat and not str(cat).startswith("="):
            category_map[str(name)] = str(cat)
    single_excluded_products = {
        norm_product(name)
        for name in [
            "配送服务费",
            "常规浓缩/浓度（ORA）",
            "加蜂蜜（ORA）",
            "加全脂奶（ORA）",
            "加手作祁红茉香茶蜜（ORA）",
            "加燕麦奶（ORA）",
            "加长白山天然椴树蜜",
            "加燕麦奶",
            "加牛奶",
            "加手作祁红茉香茶蜜",
            "抹茶加份（ORA）",
            "浓缩加份（ORA）",
            "浓缩减份（ORA）",
            "咸芝士奶盖（ORA）",
            "咸芝士奶盖（分装）",
        ]
    }

    def load_ora_product_table() -> pd.DataFrame:
        return read_excel_columns(
            "Ora外送商品数据.xlsx",
            {
                "date_id": ["date_id"],
                "sku_name": ["sku_name"],
                "quantity": ["quantity"],
                "gross_amount": ["gross_amount"],
            },
        )

    def single_aggregate_from_ora_product() -> pd.DataFrame:
        df = period_rows(load_ora_product_table(), "date_id", START, END)
        df = df[df["sku_name"].notna()].copy()
        df["_name"] = df["sku_name"].astype(str).str.strip()
        df = df[(df["_name"] != "") & (df["_name"].str.lower() != "nan")].copy()
        df["_qty"] = to_num(df["quantity"])
        df["_sales"] = to_num(df["gross_amount"])
        df = df[(df["_qty"] > 0) & (df["_sales"] > 0)].copy()
        df = df[~df["_name"].apply(lambda x: norm_product(x) in single_excluded_products)].copy()
        single = df[["_name", "_qty", "_sales"]]
        if single.empty:
            return pd.DataFrame(columns=["_name", "qty", "sales"])
        out = single.groupby("_name", as_index=False).agg(qty=("_qty", "sum"), sales=("_sales", "sum"))
        return out.sort_values(["qty", "sales"], ascending=[False, False])

    def load_product_period(start: pd.Timestamp, end: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
        mt = period_rows(
            read_excel_columns(
                "美团商品数据.xlsx",
                {
                    "日期": ["日期"],
                    "商品名": ["商品名", "商品名称"],
                    "商品销量": ["商品销量", "销量"],
                    "商品销售额": ["商品销售额", "销售额"],
                },
            ),
            "日期",
            start,
            end,
        )
        mt["_name"] = mt["商品名"].astype(str).str.strip()
        mt["_qty"] = to_num(mt["商品销量"])
        mt["_sales"] = to_num(mt["商品销售额"])
        mt = mt[(mt["_qty"] > 0) & (mt["_sales"] > 0)].copy()
        mt["_is_package"] = mt["_name"].apply(lambda x: canonical_package(x) is not None)

        ele = period_rows(
            read_excel_columns(
                "饿了么商品数据.xlsx",
                {
                    "日期": ["日期"],
                    "商品名称": ["商品名称", "商品名"],
                    "销量": ["销量", "商品销量"],
                    "销售额": ["销售额", "商品销售额"],
                    "是否套餐": ["是否套餐"],
                    "是否配料": ["是否配料"],
                },
                sheet_name="data",
            ),
            "日期",
            start,
            end,
        )
        ele["_name"] = ele["商品名称"].astype(str).str.strip()
        ele["_qty"] = to_num(ele["销量"])
        ele["_sales"] = to_num(ele["销售额"])
        ele = ele[(ele["_qty"] > 0) & (ele["_sales"] > 0)].copy()
        ele["_is_package"] = ele["_name"].apply(lambda x: canonical_package(x) is not None)
        if "是否配料" in ele.columns:
            ele["_is_addon"] = ele["是否配料"].astype(str).eq("是")
        else:
            ele["_is_addon"] = False
        return mt, ele

    def single_aggregate(mt: pd.DataFrame, ele: pd.DataFrame) -> pd.DataFrame:
        single_parts = [
            mt[(mt["_sales"] > 0) & (~mt["_is_package"])][["_name", "_qty", "_sales"]],
            ele[(ele["_sales"] > 0) & (~ele["_is_package"]) & (~ele["_is_addon"])][["_name", "_qty", "_sales"]],
        ]
        single = pd.concat(single_parts, ignore_index=True)
        if single.empty:
            return pd.DataFrame(columns=["_name", "qty", "sales"])
        out = single.groupby("_name", as_index=False).agg(qty=("_qty", "sum"), sales=("_sales", "sum"))
        return out.sort_values(["qty", "sales"], ascending=[False, False])

    def package_lookup(mt: pd.DataFrame, ele: pd.DataFrame) -> dict[str, dict[str, float]]:
        pkg_frames = []
        mt_pkg = mt[mt["_is_package"]].copy()
        mt_pkg["_canon"] = mt_pkg["_name"].apply(lambda x: canonical_package(x) or str(x).strip())
        pkg_frames.append(mt_pkg[["_canon", "_qty", "_sales"]])
        ele_pkg = ele[ele["_is_package"]].copy()
        ele_pkg["_canon"] = ele_pkg["_name"].apply(lambda x: canonical_package(x) or str(x).strip())
        pkg_frames.append(ele_pkg[["_canon", "_qty", "_sales"]])
        packages = pd.concat(pkg_frames, ignore_index=True)
        packages = packages[packages["_canon"].notna()]
        if packages.empty:
            return {}
        pkg_ag = packages.groupby("_canon", as_index=False).agg(qty=("_qty", "sum"), sales=("_sales", "sum"))
        return {str(r["_canon"]): {"qty": float(r["qty"]), "sales": float(r["sales"])} for _, r in pkg_ag.iterrows()}

    mt, ele = load_product_period(START, END)
    prev_mt_src, prev_ele_src = load_product_period(PREV_START, PREV_END)
    single_ag = single_aggregate_from_ora_product()

    rows: list[dict[str, Any]] = []
    denom = total_store_days or PERIOD_DAYS
    for _, row in single_ag.iterrows():
        name = str(row["_name"])
        rows.append(
            {
                "name": name,
                "category": infer_category(name, category_map),
                "qty": float(row["qty"]),
                "usd": safe_div(row["qty"], denom),
                "sales": float(row["sales"]),
            }
        )

    pkg_lookup = package_lookup(mt, ele)
    new_packages = sorted([name for name in pkg_lookup if name not in PACKAGE_CANONICAL], key=lambda n: pkg_lookup[n]["qty"], reverse=True)

    pkg_rows: list[dict[str, Any]] = []
    for name in PACKAGE_CANONICAL + new_packages:
        vals = pkg_lookup.get(name, {"qty": 0.0, "sales": 0.0})
        pkg_rows.append({"name": name, "qty": vals["qty"], "usd": safe_div(vals["qty"], denom), "sales": vals["sales"]})

    prev_single_qty: dict[str, float] = {}
    for row in range(2, prev_single_total_row):
        name = prev_single.cell(row, 3).value
        qty = prev_single.cell(row, 4).value
        if name:
            prev_single_qty[str(name)] = float(qty or 0)

    prev_pkg_qty: dict[str, float] = {}
    prev_pkg_lookup = package_lookup(prev_mt_src, prev_ele_src)
    if prev_pkg_lookup:
        prev_pkg_qty = {name: vals["qty"] for name, vals in prev_pkg_lookup.items()}
    else:
        prev_pkg = prev_wb["商品销售排行-套餐"]
        for row in range(2, 9):
            name = prev_pkg.cell(row, 2).value
            qty = prev_pkg.cell(row, 3).value
            if name:
                canon = canonical_package(name) or str(name)
                prev_pkg_qty[canon] = prev_pkg_qty.get(canon, 0.0) + float(qty or 0)

    return rows, pkg_rows, prev_single_qty, prev_pkg_qty, new_packages


def compute_complaints() -> tuple[dict[str, float], list[dict[str, Any]]]:
    categories = ["少餐具", "少菜品", "少酱料", "送错餐", "未看备注", "菜品口味", "菜品质量", "性价比低", "菜品异物", "客户原因", "物流问题", "服务态度差", "包装问题", "其他"]
    summary = read_excel("ora_评价汇总_2026-06-15~2026-06-21.xlsx", sheet_name="双平台中差评汇总", header=None)
    current_counts = {cat: 0.0 for cat in categories}
    # The sheet repeats the same category block for 双平台, 饿了么, and 美团.
    # Use only the first 14-row 双平台 block (Excel rows 7:20).
    for _, row in summary.iloc[6:20].iterrows():
        cat = row.iloc[3] if len(row) > 3 else None
        stat = row.iloc[4] if len(row) > 4 else None
        if cat in current_counts:
            current_counts[cat] += float(pd.to_numeric(stat, errors="coerce") or 0)

    data = read_excel("ora_评价汇总_2026-06-15~2026-06-21.xlsx", sheet_name="数据源")
    for col in categories:
        if col not in data:
            data[col] = 0
    grouped = data.groupby("外卖通门店名称", as_index=False)[categories].sum()
    grouped["total_bad"] = grouped[categories].sum(axis=1)
    grouped = grouped[grouped["total_bad"] > 0].sort_values("total_bad", ascending=False).head(10)
    rows = []
    for _, row in grouped.iterrows():
        rows.append({"store": row["外卖通门店名称"], "total": float(row["total_bad"]), **{cat: float(row[cat]) for cat in categories}})
    return current_counts, rows


def compute_delivery(mt_to_code: dict[str, str], ele_to_code: dict[str, str]) -> tuple[dict[str, float], dict[str, float]]:
    mt = read_excel("美团平均配送时长ora_自定义报表_2026-06-15_2026-06-21.xlsx")
    mt["_id"] = mt["平台门店ID"].map(norm_id)
    mt["code"] = mt["_id"].map(mt_to_code)
    mt_delivery = {str(code): float(to_num(group["平均配送时长"]).mean()) for code, group in mt[mt["code"].notna()].groupby("code")}

    ele = current_rows(read_excel("饿了么订单数据.xlsx"), "日期")
    ele["_id"] = ele["门店编号"].map(norm_id)
    ele["code"] = ele["_id"].map(ele_to_code)
    ele = ele[ele["code"].notna()].copy()
    if "订单状态" in ele.columns:
        ele = ele[ele["订单状态"].eq("订单完结")].copy()
    if "是否预订单" in ele.columns:
        ele = ele[ele["是否预订单"].astype(str).eq("否")].copy()
    ele["_accept"] = pd.to_datetime(ele["接单时间"], errors="coerce")
    ele["_done"] = pd.to_datetime(ele["完成时间"], errors="coerce")
    ele["_minutes"] = (ele["_done"] - ele["_accept"]).dt.total_seconds() / 60
    ele = ele[(ele["_minutes"].notna()) & (ele["_minutes"] >= 0)]
    ele_delivery = {str(code): float(group["_minutes"].mean()) for code, group in ele.groupby("code")}
    return mt_delivery, ele_delivery


def write_main_sheet(wb, prev_wb, stores: list[Store], metrics: dict[str, dict[str, Any]], totals: dict[str, Any]) -> None:
    ws = wb[CURRENT_SHEET]
    prev = previous_period_sheet(prev_wb)

    operating_row_by_code, operating_row_by_name = section_store_row_maps(ws, "营业数据", 3, 18)
    prev_operating_by_code, prev_operating_by_name = section_store_row_maps(prev, "营业数据", 3, 18)

    def prev_value(row: int | None, col: int) -> Any:
        return get_prev(prev, row, col) if row else None

    def write_operating(row: int, data: dict[str, Any], prev_row: int | None) -> None:
        total = data["total"]
        mt = data["mt"]
        ele = data["ele"]
        values = {
            4: total["sales_daily"],
            5: total["sales"],
            6: growth(total["sales"], prev_value(prev_row, 5)),
            7: total["discount"],
            8: total["discount_rate"],
            9: total["orders_daily"],
            10: total["orders"],
            11: total["at"],
            12: diff(total["at"], prev_value(prev_row, 11)),
            14: mt["sales_daily"],
            15: mt["sales"],
            16: growth(mt["sales"], prev_value(prev_row, 15)),
            17: mt["discount"],
            18: mt["discount_rate"],
            19: mt["orders_daily"],
            20: mt["orders"],
            21: mt["at"],
            22: diff(mt["at"], prev_value(prev_row, 21)),
            24: ele["sales_daily"],
            25: ele["sales"],
            26: growth(ele["sales"], prev_value(prev_row, 25)),
            27: ele["discount"],
            28: ele["discount_rate"],
            29: ele["orders_daily"],
            30: ele["orders"],
            31: ele["at"],
            32: diff(ele["at"], prev_value(prev_row, 31)),
        }
        for col, value in values.items():
            write(ws.cell(row, col), value)
        for col in [34, 35, 37, 39, 40, 42, 43, 45, 47, 48]:
            write(ws.cell(row, col), 0)

    for store in stores:
        row = matched_row_by_store(store, operating_row_by_code, operating_row_by_name)
        if row:
            prev_row = matched_row_by_store(store, prev_operating_by_code, prev_operating_by_name)
            write_operating(row, metrics[store.code], prev_row)

    total_data = {"total": totals["total"], "mt": totals["mt"], "ele": totals["ele"]}
    write_operating(19, total_data, find_section_total_row(prev, "营业数据", 19))
    for row in range(3, 20):
        for col in [9, 10, 19, 20, 29, 30, 39, 40, 48, 49]:
            ws.cell(row, col).number_format = "#,##0"

    def write_traffic(row: int, data: dict[str, Any], prev_row: int | None) -> None:
        total, mt, ele = data["total"], data["mt"], data["ele"]
        values = {
            4: total["exp_people_daily"],
            5: total["exp_people"],
            6: growth(total["exp_people"], prev_value(prev_row, 5)),
            7: total["entry_rate"],
            8: diff(total["entry_rate"], prev_value(prev_row, 7)),
            9: total["order_rate"],
            11: diff(total["order_rate"], prev_value(prev_row, 9)),
            12: total["exp_count"],
            14: mt["exp_people_daily"],
            15: mt["exp_people"],
            16: growth(mt["exp_people"], prev_value(prev_row, 15)),
            17: mt["entry_rate"],
            18: diff(mt["entry_rate"], prev_value(prev_row, 17)),
            19: mt["order_rate"],
            21: diff(mt["order_rate"], prev_value(prev_row, 19)),
            22: mt["exp_count"],
            24: ele["exp_people_daily"],
            25: ele["exp_people"],
            26: growth(ele["exp_people"], prev_value(prev_row, 25)),
            27: ele["entry_rate"],
            28: diff(ele["entry_rate"], prev_value(prev_row, 27)),
            29: ele["order_rate"],
            31: diff(ele["order_rate"], prev_value(prev_row, 29)),
            32: ele["exp_count"],
        }
        for col, value in values.items():
            write(ws.cell(row, col), value)

    traffic_row_by_code, traffic_row_by_name = section_store_row_maps(ws, "流量数据", 23, 38)
    prev_traffic_by_code, prev_traffic_by_name = section_store_row_maps(prev, "流量数据", 23, 38)
    for store in stores:
        row = matched_row_by_store(store, traffic_row_by_code, traffic_row_by_name)
        if row:
            prev_row = matched_row_by_store(store, prev_traffic_by_code, prev_traffic_by_name)
            write_traffic(row, metrics[store.code], prev_row)
    write_traffic(39, total_data, find_section_total_row(prev, "流量数据", 39))
    for row in range(23, 40):
        for col in [4, 5, 12, 14, 15, 22, 24, 25, 32]:
            ws.cell(row, col).number_format = "#,##0"

    def write_promo(row: int, data: dict[str, Any]) -> None:
        total, mt, ele = data["total"], data["mt"], data["ele"]
        values = {
            4: total["paid_exp"],
            6: total["ad_share"],
            7: total["ad_spend"],
            8: total["ad_visits"],
            9: total["ad_roi"],
            10: total["ad_orig"],
            11: total["ad_gmv_share"],
            14: mt["paid_exp"],
            16: mt["ad_share"],
            17: mt["ad_spend"],
            18: mt["ad_visits"],
            19: mt["ad_roi"],
            20: mt["ad_orig"],
            21: mt["ad_gmv_share"],
            24: ele["paid_exp"],
            26: ele["ad_share"],
            27: ele["ad_spend"],
            28: ele["ad_visits"],
            29: ele["ad_roi"],
            30: ele["ad_orig"],
            31: ele["ad_gmv_share"],
        }
        for col, value in values.items():
            write(ws.cell(row, col), value)

    promo_row_by_code, promo_row_by_name = section_store_row_maps(ws, "推广数据", 43, 58)
    for store in stores:
        row = matched_row_by_store(store, promo_row_by_code, promo_row_by_name)
        if row:
            write_promo(row, metrics[store.code])
    write_promo(59, total_data)
    for row in range(43, 60):
        for col in [6, 11, 16, 21, 26, 31]:
            ws.cell(row, col).number_format = "0.0%"
        for col in [4, 8, 10, 14, 18, 20, 24, 28, 30]:
            ws.cell(row, col).number_format = "#,##0"

    score_row_by_code, score_row_by_name = section_store_row_maps(ws, "门店评分", 63, 78)
    for store in stores:
        row = matched_row_by_store(store, score_row_by_code, score_row_by_name)
        if not row:
            continue
        data = metrics[store.code]
        values = {
            4: data["total"]["score"],
            6: data["total"]["good"],
            7: data["total"]["bad"],
            14: data["mt"]["score"],
            16: data["mt"]["good"],
            17: data["mt"]["bad"],
            24: data["ele"]["score"],
            26: data["ele"]["good"],
            27: data["ele"]["bad"],
        }
        for col, value in values.items():
            write(ws.cell(row, col), value)


ANALYSIS_SECTION_FALLBACKS = {
    "营业数据": (3, 18, 19),
    "流量数据": (23, 38, 39),
    "推广数据": (43, 58, 59),
    "门店评分": (63, 78, 79),
}


def norm_header(value: Any) -> str:
    text = cell_text(value).replace("_", "")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def header_cols(
    sheet,
    header_row: int,
    candidates: list[str],
    start_col: int = 1,
    end_col: int | None = None,
    exact: bool = True,
) -> list[int]:
    end = min(end_col or sheet.max_column or 1, sheet.max_column or 1)
    normalized = [norm_header(candidate) for candidate in candidates]
    matches: list[int] = []
    for col in range(max(1, start_col), end + 1):
        text = norm_header(sheet.cell(header_row, col).value)
        if not text:
            continue
        for candidate in normalized:
            if (exact and text == candidate) or ((not exact) and candidate in text):
                matches.append(col)
                break
    return matches


def first_header_col(
    sheet,
    header_row: int,
    candidates: list[str],
    start_col: int = 1,
    end_col: int | None = None,
    exact: bool = True,
) -> int | None:
    cols = header_cols(sheet, header_row, candidates, start_col, end_col, exact)
    return cols[0] if cols else None


def nth_header_col(sheet, header_row: int, candidates: list[str], nth: int, exact: bool = False) -> int | None:
    cols = header_cols(sheet, header_row, candidates, 1, None, exact)
    return cols[nth - 1] if len(cols) >= nth else None


def scoped_header_col(
    sheet,
    header_row: int,
    scope_ranges: dict[str, tuple[int, int]],
    scope: str,
    candidates: list[str],
    exact: bool = True,
) -> int | None:
    start_col, end_col = scope_ranges.get(scope, (1, sheet.max_column or 1))
    return first_header_col(sheet, header_row, candidates, start_col, end_col, exact)


def section_header(sheet, title: str) -> int:
    start, _end, _total = ANALYSIS_SECTION_FALLBACKS[title]
    return find_section_header_row(sheet, title, start)


def section_total(sheet, title: str) -> int:
    _start, _end, total = ANALYSIS_SECTION_FALLBACKS[title]
    return find_section_total_row(sheet, title, total)


def section_rows_for_analysis(sheet, title: str) -> tuple[dict[str, int], dict[str, int]]:
    start, end, _total = ANALYSIS_SECTION_FALLBACKS[title]
    return section_store_row_maps(sheet, title, start, end)


def analysis_row_maps(sheet) -> dict[str, tuple[dict[str, int], dict[str, int]]]:
    return {title: section_rows_for_analysis(sheet, title) for title in ("营业数据", "流量数据", "推广数据", "门店评分")}


def find_label_row(sheet, label: str, col: int | None = 1, start: int = 1, end: int | None = None) -> int | None:
    last = min(end or sheet.max_row or 1, sheet.max_row or 1)
    for row in range(start, last + 1):
        if col is not None and cell_text(sheet.cell(row, col).value) == label:
            return row
        if col is None:
            for cur_col in range(1, min(sheet.max_column or 1, 20) + 1):
                if cell_text(sheet.cell(row, cur_col).value) == label:
                    return row
    return None


def find_label_cell(sheet, label: str, start: int = 1, end: int | None = None) -> tuple[int, int] | None:
    last = min(end or sheet.max_row or 1, sheet.max_row or 1)
    for row in range(start, last + 1):
        for col in range(1, min(sheet.max_column or 1, 20) + 1):
            if cell_text(sheet.cell(row, col).value) == label:
                return row, col
    return None


def analysis_scope_ranges(anchor_cols: dict[str, int | None], max_col: int) -> dict[str, tuple[int, int]]:
    mt = anchor_cols.get("mt")
    ele = anchor_cols.get("ele")
    return {
        "total": (1, (mt or max_col + 1) - 1),
        "mt": (mt or 1, (ele or max_col + 1) - 1),
        "ele": (ele or 1, max_col),
    }


def resolve_analysis_schema(sheet) -> dict[str, dict[str, dict[str, int | None]]]:
    op_header = section_header(sheet, "营业数据")
    traffic_header = section_header(sheet, "流量数据")
    promo_header = section_header(sheet, "推广数据")
    score_header = section_header(sheet, "门店评分")
    max_col = sheet.max_column or 1

    op_anchors = {
        "total": first_header_col(sheet, op_header, ["总sales"], exact=True),
        "mt": first_header_col(sheet, op_header, ["美团Sales"], exact=True),
        "ele": first_header_col(sheet, op_header, ["饿了么Sales"], exact=True),
    }
    op_ranges = analysis_scope_ranges(op_anchors, max_col)

    traffic_anchors = {
        "total": first_header_col(sheet, traffic_header, ["总曝光人数"], exact=True),
        "mt": first_header_col(sheet, traffic_header, ["美团曝光人数"], exact=True),
        "ele": first_header_col(sheet, traffic_header, ["饿了么曝光人数"], exact=True),
    }
    traffic_ranges = analysis_scope_ranges(traffic_anchors, max_col)

    schema: dict[str, dict[str, dict[str, int | None]]] = {
        "营业数据": {},
        "流量数据": {},
        "推广数据": {},
        "门店评分": {},
    }
    for scope in ("total", "mt", "ele"):
        schema["营业数据"][scope] = {
            "sales": op_anchors[scope],
            "discount_rate": scoped_header_col(sheet, op_header, op_ranges, scope, ["商户折扣率"], exact=True),
            "orders": scoped_header_col(sheet, op_header, op_ranges, scope, ["ADT"], exact=True),
            "at": scoped_header_col(sheet, op_header, op_ranges, scope, ["AT"], exact=True),
        }
        schema["流量数据"][scope] = {
            "exp_people_daily": scoped_header_col(sheet, traffic_header, traffic_ranges, scope, ["曝光人数日均", "曝光人数_日均", "曝光人数 日均"], exact=False),
            "exp_people": traffic_anchors[scope],
            "entry_rate": scoped_header_col(sheet, traffic_header, traffic_ranges, scope, ["进店转化率"], exact=True),
            "order_rate": scoped_header_col(sheet, traffic_header, traffic_ranges, scope, ["下单转化率"], exact=True),
            "exp_count": scoped_header_col(
                sheet,
                traffic_header,
                traffic_ranges,
                scope,
                {
                    "total": ["总曝光次数"],
                    "mt": ["美团曝光次数"],
                    "ele": ["饿了么曝光次数"],
                }[scope],
                exact=True,
            ),
        }

    promo_occurrence = {"total": 1, "mt": 2, "ele": 3}
    for scope, occurrence in promo_occurrence.items():
        schema["推广数据"][scope] = {
            "paid_exp": nth_header_col(sheet, promo_header, ["曝光次数"], occurrence, exact=False),
            "ad_spend": nth_header_col(sheet, promo_header, ["消耗金额"], occurrence, exact=False),
            "ad_visits": nth_header_col(sheet, promo_header, ["进店数"], occurrence, exact=False),
            "ad_roi": nth_header_col(sheet, promo_header, ["营业额ROI"], occurrence, exact=False),
            "ad_orig": nth_header_col(sheet, promo_header, ["订单原价交易额"], occurrence, exact=False),
        }

    schema["门店评分"]["total"] = {
        "bad": scoped_header_col(sheet, score_header, {"total": (1, max_col)}, "total", ["中差评"], exact=True),
    }
    return schema


def analysis_cell_num(sheet, row: int | None, col: int | None) -> float:
    if not row or not col:
        return 0.0
    return scalar_num(clean_error_value(sheet.cell(row, col).value))


def read_analysis_record(
    sheet,
    schema: dict[str, dict[str, dict[str, int | None]]],
    rows: dict[str, int | None],
    scope: str,
) -> dict[str, float]:
    op = rows.get("营业数据")
    traffic = rows.get("流量数据")
    promo = rows.get("推广数据")
    score = rows.get("门店评分")
    op_cols = schema["营业数据"].get(scope, {})
    traffic_cols = schema["流量数据"].get(scope, {})
    promo_cols = schema["推广数据"].get(scope, {})
    score_cols = schema["门店评分"].get("total", {})
    return {
        "sales": analysis_cell_num(sheet, op, op_cols.get("sales")),
        "discount_rate": analysis_cell_num(sheet, op, op_cols.get("discount_rate")),
        "orders": analysis_cell_num(sheet, op, op_cols.get("orders")),
        "at": analysis_cell_num(sheet, op, op_cols.get("at")),
        "exp_people_daily": analysis_cell_num(sheet, traffic, traffic_cols.get("exp_people_daily")),
        "exp_people": analysis_cell_num(sheet, traffic, traffic_cols.get("exp_people")),
        "entry_rate": analysis_cell_num(sheet, traffic, traffic_cols.get("entry_rate")),
        "order_rate": analysis_cell_num(sheet, traffic, traffic_cols.get("order_rate")),
        "exp_count": analysis_cell_num(sheet, traffic, traffic_cols.get("exp_count")),
        "paid_exp": analysis_cell_num(sheet, promo, promo_cols.get("paid_exp")),
        "ad_spend": analysis_cell_num(sheet, promo, promo_cols.get("ad_spend")),
        "ad_visits": analysis_cell_num(sheet, promo, promo_cols.get("ad_visits")),
        "ad_roi": analysis_cell_num(sheet, promo, promo_cols.get("ad_roi")),
        "ad_orig": analysis_cell_num(sheet, promo, promo_cols.get("ad_orig")),
        "bad": analysis_cell_num(sheet, score, score_cols.get("bad")),
    }


def record_delta(cur: dict[str, float], prev: dict[str, float], key: str) -> float:
    return scalar_num(cur.get(key, 0.0)) - scalar_num(prev.get(key, 0.0))


def record_growth(cur: dict[str, float], prev: dict[str, float], key: str) -> float | None:
    return growth(cur.get(key, 0.0), prev.get(key, 0.0))


def trend_word(value: float, positive: str = "增长", negative: str = "下滑", zero: str = "持平") -> str:
    if value > 1e-9:
        return positive
    if value < -1e-9:
        return negative
    return zero


def fmt_int_abs(value: float) -> str:
    return f"{int(round(abs(value))):,}"


def fmt_signed_int(value: float, unit: str = "") -> str:
    number = int(round(value))
    if number > 0:
        return f"+{number:,}{unit}"
    if number < 0:
        return f"-{abs(number):,}{unit}"
    return f"0{unit}"


def fmt_signed_int_with_hold(value: float, unit: str = "") -> str:
    if int(round(value)) == 0:
        return f"持平0{unit}"
    return fmt_signed_int(value, unit)


def fmt_signed_float(value: float, unit: str = "") -> str:
    if value > 1e-9:
        return f"+{value:.1f}{unit}"
    if value < -1e-9:
        return f"-{abs(value):.1f}{unit}"
    return f"0.0{unit}"


def fmt_signed_float_with_hold(value: float, unit: str = "") -> str:
    if abs(round(value, 1)) <= 0:
        return f"持平0.0{unit}"
    return fmt_signed_float(value, unit)


def fmt_pct_abs(value: float | None) -> str:
    if value is None:
        return "0.0%"
    return f"{abs(value) * 100:.1f}%"


def fmt_signed_pct(value: float | None) -> str:
    if value is None or abs(round(value * 100, 1)) <= 0:
        return "持平0.0%"
    sign = "+" if value > 0 else "-"
    return f"{sign}{abs(value) * 100:.1f}%"


def fmt_level_pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def fmt_sales_change(label: str, cur: float, prev: float) -> str:
    delta = cur - prev
    gr = growth(cur, prev)
    if int(round(delta)) == 0 and (gr is None or abs(round(gr * 100, 1)) <= 0):
        return f"{label}持平0.0%（0元）"
    word = trend_word(delta)
    if gr is None:
        return f"{label}{word}0.0%（{fmt_signed_int(delta, '元')}）"
    return f"{label}{word}{fmt_pct_abs(gr)}（{fmt_signed_int(delta, '元')}）"


def fmt_sales_or_orders(label: str, cur: float, prev: float, unit: str) -> str:
    if unit == "元":
        return fmt_sales_change(label, cur, prev)
    delta = cur - prev
    if int(round(delta)) == 0:
        return f"{label}持平0{unit}"
    return f"{label}{fmt_signed_int(delta, unit)}"


def fmt_at_delta(cur: float, prev: float, label: str = "AT") -> str:
    delta = cur - prev
    return f"{label}{fmt_signed_float_with_hold(delta, '元')}"


def fmt_order_delta(cur: float, prev: float) -> str:
    delta = cur - prev
    if int(round(delta)) == 0:
        return "订单持平0单"
    return f"订单{fmt_signed_int(delta, '单')}"


def fmt_platform_sales(label: str, cur: float, prev: float) -> str:
    delta = cur - prev
    gr = growth(cur, prev)
    if int(round(delta)) == 0 and (gr is None or abs(round(gr * 100, 1)) <= 0):
        return f"{label}sales持平0.0%"
    if gr is None:
        return f"{label}sales{trend_word(delta)}0.0%"
    return f"{label}sales{trend_word(delta)}{fmt_pct_abs(gr)}"


def fmt_rate_delta(label: str, cur: float, prev: float) -> str:
    delta = cur - prev
    if abs(round(delta * 100, 1)) <= 0:
        return f"{label}持平"
    return f"{label}{trend_word(delta)}{abs(delta) * 100:.1f}%"


def same_direction(metric_delta: float, sales_delta: float) -> bool:
    if sales_delta > 1e-9:
        return metric_delta > 1e-9
    if sales_delta < -1e-9:
        return metric_delta < -1e-9
    return abs(metric_delta) <= 1e-9


def include_platform_factor(metric_delta: float, sales_delta: float) -> bool:
    if abs(sales_delta) <= 1e-9:
        return abs(metric_delta) <= 1e-9
    return same_direction(metric_delta, sales_delta) and abs(metric_delta) > 1e-9


def trend_bucket(value: float, zero_threshold: float) -> int:
    if value > zero_threshold:
        return 1
    if value < -zero_threshold:
        return -1
    return 0


def include_platform_factor_display(metric_delta: float, sales_delta: float, zero_threshold: float) -> bool:
    return trend_bucket(metric_delta, zero_threshold) == trend_bucket(sales_delta, 1e-9)


def exposure_phrase(cur: dict[str, float], prev: dict[str, float]) -> str:
    exp_growth = record_growth(cur, prev, "exp_people")
    exp_delta = record_delta(cur, prev, "exp_people")
    display_delta = exp_delta
    if exp_growth is not None and abs(round(exp_growth * 100, 1)) <= 0:
        display_delta = 0.0
    word = trend_word(display_delta)
    phrase = f"曝光量{word}{fmt_pct_abs(exp_growth)}"
    if display_delta >= -1e-9:
        return phrase

    total_growth = record_growth(cur, prev, "exp_count")
    cur_natural = max(cur.get("exp_count", 0.0) - cur.get("paid_exp", 0.0), 0.0)
    prev_natural = max(prev.get("exp_count", 0.0) - prev.get("paid_exp", 0.0), 0.0)
    natural_growth = growth(cur_natural, prev_natural)
    paid_growth = record_growth(cur, prev, "paid_exp")
    pieces = [
        f"总曝光次数{trend_word(record_delta(cur, prev, 'exp_count'))}{fmt_pct_abs(total_growth)}",
        f"自然曝光次数{trend_word(cur_natural - prev_natural)}{fmt_pct_abs(natural_growth)}",
        f"付费曝光次数{trend_word(record_delta(cur, prev, 'paid_exp'))}{fmt_pct_abs(paid_growth)}",
    ]
    total_drop = prev.get("exp_count", 0.0) - cur.get("exp_count", 0.0)
    paid_drop = prev.get("paid_exp", 0.0) - cur.get("paid_exp", 0.0)
    natural_drop = prev_natural - cur_natural
    if total_drop > 0 and paid_drop > 0 and natural_drop > 0:
        share = safe_div(paid_drop, total_drop)
        if share is not None:
            pieces.append(f"付费曝光占总曝光次数下滑的{fmt_pct_abs(share)}")
    return phrase + "（" + "，".join(pieces) + "）"


def build_business_scope_sentence(label: str, cur: dict[str, float], prev: dict[str, float], include_daily_sales: bool) -> str | None:
    sales_delta = record_delta(cur, prev, "sales")
    if abs(cur.get("sales", 0.0)) <= 1e-9 and abs(prev.get("sales", 0.0)) <= 1e-9:
        return None

    sales_label = f"{CURRENT_SHEET}总sales" if label == "整体" else "总sales"
    pieces = [fmt_sales_or_orders(sales_label, cur.get("sales", 0.0), prev.get("sales", 0.0), "元")]
    if include_daily_sales:
        cur_daily = safe_div(cur.get("sales", 0.0), PERIOD_DAYS) or 0.0
        prev_days = int((PREV_END - PREV_START).days) + 1
        prev_daily = safe_div(prev.get("sales", 0.0), prev_days) or 0.0
        pieces.append(fmt_sales_or_orders("日均sales", cur_daily, prev_daily, "元"))

    pieces.append(fmt_order_delta(cur.get("orders", 0.0), prev.get("orders", 0.0)))
    pieces.append(fmt_at_delta(cur.get("at", 0.0), prev.get("at", 0.0), "总AT" if label == "整体" else "AT"))

    return f"{label}： " + "，".join(pieces)


def build_platform_scope_sentence(label: str, cur: dict[str, float], prev: dict[str, float]) -> str | None:
    sales_delta = record_delta(cur, prev, "sales")
    if abs(cur.get("sales", 0.0)) <= 1e-9 and abs(prev.get("sales", 0.0)) <= 1e-9:
        return None

    reasons: list[str] = []
    exp_delta = record_delta(cur, prev, "exp_people")
    exp_growth = record_growth(cur, prev, "exp_people")
    if (
        (exp_growth is not None and include_platform_factor_display(exp_growth, sales_delta, 0.0005))
        or (exp_growth is None and include_platform_factor(exp_delta, sales_delta))
    ):
        reasons.append(exposure_phrase(cur, prev))

    entry_delta = record_delta(cur, prev, "entry_rate")
    if include_platform_factor_display(entry_delta, sales_delta, 0.0005):
        reasons.append(fmt_rate_delta("P1", cur.get("entry_rate", 0.0), prev.get("entry_rate", 0.0)))

    order_rate_delta = record_delta(cur, prev, "order_rate")
    if include_platform_factor_display(order_rate_delta, sales_delta, 0.0005):
        reasons.append(fmt_rate_delta("P2", cur.get("order_rate", 0.0), prev.get("order_rate", 0.0)))

    at_delta = record_delta(cur, prev, "at")
    if include_platform_factor_display(at_delta, sales_delta, 0.05):
        reasons.append(fmt_at_delta(cur.get("at", 0.0), prev.get("at", 0.0)))

    sentence = fmt_platform_sales(label, cur.get("sales", 0.0), prev.get("sales", 0.0))
    if reasons:
        sentence += "，主要是因为" + "，".join(reasons)
    return sentence


def business_entity_text(
    label: str,
    cur_records: dict[str, dict[str, float]],
    prev_records: dict[str, dict[str, float]],
    include_daily_sales: bool,
) -> str | None:
    total_cur = cur_records["total"]
    total_prev = prev_records["total"]
    total_sales_delta = record_delta(total_cur, total_prev, "sales")
    main = build_business_scope_sentence(label, total_cur, total_prev, include_daily_sales)
    if not main:
        return None

    platform_sentences: list[str] = []
    for scope, name in [("mt", "美团"), ("ele", "饿了么")]:
        platform_label = f"{name}整体" if label == "整体" else name
        sentence = build_platform_scope_sentence(platform_label, cur_records[scope], prev_records[scope])
        if sentence:
            platform_sentences.append(sentence)
    if platform_sentences:
        return main + "。" + "；".join(platform_sentences) + "。"
    return main + "。"


def analysis_rows_for_store(
    sheet,
    code: Any,
    name: Any,
    row_maps: dict[str, tuple[dict[str, int], dict[str, int]]] | None = None,
) -> dict[str, int | None]:
    rows: dict[str, int | None] = {}
    maps_by_title = row_maps or analysis_row_maps(sheet)
    for title in ("营业数据", "流量数据", "推广数据", "门店评分"):
        rows[title] = matched_row_by_code_name(code, name, *maps_by_title[title])
    return rows


def analysis_total_rows(sheet) -> dict[str, int | None]:
    return {title: section_total(sheet, title) for title in ("营业数据", "流量数据", "推广数据", "门店评分")}


def read_records_for_rows(sheet, schema: dict[str, dict[str, dict[str, int | None]]], rows: dict[str, int | None]) -> dict[str, dict[str, float]]:
    return {scope: read_analysis_record(sheet, schema, rows, scope) for scope in ("total", "mt", "ele")}


def build_business_analysis_text(wb) -> str:
    cur_ws = wb[CURRENT_SHEET]
    prev_ws = wb["上期"]
    cur_schema = resolve_analysis_schema(cur_ws)
    prev_schema = resolve_analysis_schema(prev_ws)
    cur_row_maps = analysis_row_maps(cur_ws)
    prev_row_maps = analysis_row_maps(prev_ws)
    include_daily_sales = PERIOD_DAYS != int((PREV_END - PREV_START).days) + 1

    paragraphs: list[str] = []
    total_text = business_entity_text(
        "整体",
        read_records_for_rows(cur_ws, cur_schema, analysis_total_rows(cur_ws)),
        read_records_for_rows(prev_ws, prev_schema, analysis_total_rows(prev_ws)),
        include_daily_sales,
    )
    if total_text:
        paragraphs.append(total_text)

    prev_op_maps = prev_row_maps["营业数据"]
    header_row = section_header(cur_ws, "营业数据")
    total_row = section_total(cur_ws, "营业数据")
    for row in range(header_row + 1, total_row):
        code = cur_ws.cell(row, 2).value
        name = cur_ws.cell(row, 3).value
        if not code and not name:
            continue
        prev_row = matched_row_by_code_name(code, name, *prev_op_maps)
        cur_rows = analysis_rows_for_store(cur_ws, code, name, cur_row_maps)
        prev_rows = analysis_rows_for_store(prev_ws, code, name, prev_row_maps)
        if not prev_row or not cur_rows.get("营业数据"):
            continue
        cur_records = read_records_for_rows(cur_ws, cur_schema, cur_rows)
        prev_records = read_records_for_rows(prev_ws, prev_schema, prev_rows)
        store_label = cell_text(name) or cell_text(code)
        text = business_entity_text(store_label, cur_records, prev_records, include_daily_sales)
        if text:
            paragraphs.append(text)

    return "\n\n".join(paragraphs)


def estimate_text_rows(text: str, chars_per_row: int = 58) -> int:
    rows = 0
    for line in text.splitlines() or [""]:
        rows += max(1, math.ceil(len(line) / chars_per_row))
    return max(20, rows + 3)


def write_merged_text_area(ws, text: str, start_row: int, start_col: int, end_col: int, min_rows: int = 20) -> None:
    needed_rows = max(min_rows, estimate_text_rows(text))
    end_row = start_row + needed_rows - 1
    for merged in list(ws.merged_cells.ranges):
        if not (merged.max_row < start_row or merged.min_row > end_row or merged.max_col < start_col or merged.min_col > end_col):
            ws.unmerge_cells(str(merged))
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = 24
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).value = None
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(start_row, start_col)
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")


def write_business_analysis(wb) -> None:
    ws = wb[CURRENT_SHEET]
    text = build_business_analysis_text(wb)
    if not text:
        text = "整体：本期与上期核心经营数据暂无可比变化。"
    score_total = section_total(ws, "门店评分")
    start_row = max(score_total + 2, 81)
    write_merged_text_area(ws, text, start_row, 4, 12, min_rows=20)


def email_change_word(value: float) -> str:
    if value > 1e-9:
        return "上涨"
    if value < -1e-9:
        return "下滑"
    return "持平"


def email_pct_phrase(value: float | None) -> str:
    if value is None or abs(round(value * 100, 1)) <= 0:
        return "持平0.0%"
    return f"{email_change_word(value)}{fmt_pct_abs(value)}"


def current_score_bad_total(sheet) -> float:
    score_header = section_header(sheet, "门店评分")
    bad_col = scoped_header_col(sheet, score_header, {"total": (1, sheet.max_column or 1)}, "total", ["中差评"], exact=True)
    if not bad_col:
        return 0.0
    by_code, by_name = section_rows_for_analysis(sheet, "门店评分")
    rows = sorted(set(by_code.values()) | set(by_name.values()))
    return sum(analysis_cell_num(sheet, row, bad_col) for row in rows)


def build_performance_email_text(wb) -> str:
    cur_ws = wb[CURRENT_SHEET]
    prev_ws = wb["上期"]
    cur_schema = resolve_analysis_schema(cur_ws)
    prev_schema = resolve_analysis_schema(prev_ws)
    cur_records = read_records_for_rows(cur_ws, cur_schema, analysis_total_rows(cur_ws))
    prev_records = read_records_for_rows(prev_ws, prev_schema, analysis_total_rows(prev_ws))

    cur_total = cur_records["total"]
    prev_total = prev_records["total"]
    cur_mt = cur_records["mt"]
    prev_mt = prev_records["mt"]
    cur_ele = cur_records["ele"]
    prev_ele = prev_records["ele"]

    sales_growth = record_growth(cur_total, prev_total, "sales")
    orders_delta = record_delta(cur_total, prev_total, "orders")
    at_delta = record_delta(cur_total, prev_total, "at")
    mt_sales_growth = record_growth(cur_mt, prev_mt, "sales")
    mt_sales_delta = record_delta(cur_mt, prev_mt, "sales")
    ele_sales_growth = record_growth(cur_ele, prev_ele, "sales")
    ele_sales_delta = record_delta(cur_ele, prev_ele, "sales")
    discount_delta = record_delta(cur_total, prev_total, "discount_rate")
    exp_daily_growth = record_growth(cur_total, prev_total, "exp_people_daily")
    entry_delta = record_delta(cur_total, prev_total, "entry_rate")
    order_rate_delta = record_delta(cur_total, prev_total, "order_rate")
    bad_total = current_score_bad_total(cur_ws)

    return "\n".join(
        [
            f"1、上周业绩：sales达成{fmt_int_abs(cur_total.get('sales', 0.0))}元，环比上周{email_pct_phrase(sales_growth)}；有效单{fmt_signed_int_with_hold(orders_delta, '单')}，AT{fmt_signed_float_with_hold(at_delta, '元')}",
            f"渠道表现：美团sales环比{email_pct_phrase(mt_sales_growth)}（{fmt_signed_int_with_hold(mt_sales_delta, '元')}），饿了么sales环比{email_pct_phrase(ele_sales_growth)}（{fmt_signed_int_with_hold(ele_sales_delta, '元')}）",
            f"2、折扣情况：整体折扣率为{fmt_level_pct(cur_total.get('discount_rate', 0.0))}（环比{fmt_signed_pct(discount_delta)}）",
            f"3、流量表现：店均日曝光人数{fmt_int_abs(cur_total.get('exp_people_daily', 0.0))}（环比{fmt_signed_pct(exp_daily_growth)}），进店转化率{fmt_level_pct(cur_total.get('entry_rate', 0.0))}（环比{fmt_signed_pct(entry_delta)}），下单转化率{fmt_level_pct(cur_total.get('order_rate', 0.0))}（环比{fmt_signed_pct(order_rate_delta)}）",
            f"4、推广表现：消耗金额{fmt_int_abs(cur_total.get('ad_spend', 0.0))}元，进店数{fmt_int_abs(cur_total.get('ad_visits', 0.0))}人次，预估带来sales {fmt_int_abs(cur_total.get('ad_orig', 0.0))}元，平均ROI为{cur_total.get('ad_roi', 0.0):.1f}",
            "5、服务指标",
            f"中差评：{fmt_int_abs(bad_total)}条",
        ]
    )


def write_email_content_sheet(wb) -> None:
    if "邮件内容" in wb.sheetnames:
        ws = wb["邮件内容"]
    else:
        ws = wb.create_sheet("邮件内容")
    if wb.sheetnames[-1] != "邮件内容":
        wb.move_sheet(ws, offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("邮件内容"))
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    text = build_performance_email_text(wb)
    write_merged_text_area(ws, text, 1, 1, 10, min_rows=12)


def write_v2(wb) -> None:
    ws = wb["V2"]
    sheet = CURRENT_SHEET
    cur_ws = wb[sheet]
    prev_ws = wb["上期"]
    cur_operating_total = find_section_total_row(cur_ws, "营业数据", 19)
    prev_operating_total = find_section_total_row(prev_ws, "营业数据", 19)
    cur_traffic_total = find_section_total_row(cur_ws, "流量数据", 39)
    prev_traffic_total = find_section_total_row(prev_ws, "流量数据", 39)
    cur_promo_total = find_section_total_row(cur_ws, "推广数据", 59)
    cur_section_maps = {
        "营业数据": section_store_row_maps(cur_ws, "营业数据", 3, 18, 2, 3),
        "流量数据": section_store_row_maps(cur_ws, "流量数据", 23, 38, 2, 3),
        "推广数据": section_store_row_maps(cur_ws, "推广数据", 43, 58, 2, 3),
        "门店评分": section_store_row_maps(cur_ws, "门店评分", 63, 78, 2, 3),
    }
    prev_section_maps = {
        "营业数据": section_store_row_maps(prev_ws, "营业数据", 3, 18, 2, 3),
        "流量数据": section_store_row_maps(prev_ws, "流量数据", 23, 38, 2, 3),
        "推广数据": section_store_row_maps(prev_ws, "推广数据", 43, 58, 2, 3),
        "门店评分": section_store_row_maps(prev_ws, "门店评分", 63, 78, 2, 3),
    }

    def section_rows(title: str, cur_start: int, prev_start: int, v2_row: int, offset: int, cur_fallback: int) -> tuple[int, int | None]:
        code = ws.cell(v2_row, 1).value
        name = ws.cell(v2_row, 2).value
        cur_row = matched_row_by_code_name(code, name, *cur_section_maps[title]) or cur_fallback
        prev_row = matched_row_by_code_name(code, name, *prev_section_maps[title])
        return cur_row, prev_row

    for i in range(16):
        r_v2 = 3 + i
        r_cur, _ = section_rows("营业数据", 3, 3, r_v2, i, 3 + i)
        ws.cell(r_v2, 3).value = f"='{sheet}'!D{r_cur}"
        ws.cell(r_v2, 4).value = f"='{sheet}'!F{r_cur}"
        ws.cell(r_v2, 5).value = f"='{sheet}'!P{r_cur}"
        ws.cell(r_v2, 6).value = f"='{sheet}'!Z{r_cur}"
    ws.cell(19, 3).value = f"='{sheet}'!D{cur_operating_total}"
    ws.cell(19, 4).value = f"='{sheet}'!F{cur_operating_total}"
    ws.cell(19, 5).value = f"='{sheet}'!P{cur_operating_total}"
    ws.cell(19, 6).value = f"='{sheet}'!Z{cur_operating_total}"

    for i in range(16):
        r_v2 = 23 + i
        r_cur, r_prev = section_rows("营业数据", 3, 3, r_v2, i, 3 + i)
        ws.cell(r_v2, 3).value = f"='{sheet}'!K{r_cur}"
        ws.cell(r_v2, 4).value = f"='{sheet}'!L{r_cur}"
        ws.cell(r_v2, 5).value = f"='{sheet}'!H{r_cur}"
        ws.cell(r_v2, 6).value = f"=E{r_v2}-H{r_v2}"
        ws.cell(r_v2, 8).value = f"='上期'!H{r_prev}" if r_prev else None
    ws.cell(39, 3).value = f"='{sheet}'!K{cur_operating_total}"
    ws.cell(39, 4).value = f"='{sheet}'!L{cur_operating_total}"
    ws.cell(39, 5).value = f"='{sheet}'!H{cur_operating_total}"
    ws.cell(39, 6).value = "=E39-H39"
    ws.cell(39, 8).value = f"='上期'!H{prev_operating_total}"

    for i in range(16):
        r_v2 = 43 + i
        r_cur, r_prev = section_rows("流量数据", 23, 23, r_v2, i, 23 + i)
        for c_v2, c_cur in zip([3, 5, 6, 7, 8], ["D", "G", "H", "I", "K"]):
            ws.cell(r_v2, c_v2).value = f"='{sheet}'!{c_cur}{r_cur}"
        ws.cell(r_v2, 4).value = f"=IFERROR('{sheet}'!D{r_cur}/'上期'!D{r_prev}-1,0)" if r_prev else None
        ws.cell(r_v2, 4).number_format = "0.0%"
    ws.cell(59, 3).value = f"='{sheet}'!D{cur_traffic_total}"
    ws.cell(59, 4).value = f"=IFERROR('{sheet}'!D{cur_traffic_total}/'上期'!D{prev_traffic_total}-1,0)"
    ws.cell(59, 4).number_format = "0.0%"
    ws.cell(59, 5).value = f"='{sheet}'!G{cur_traffic_total}"
    ws.cell(59, 6).value = f"='{sheet}'!H{cur_traffic_total}"
    ws.cell(59, 7).value = f"='{sheet}'!I{cur_traffic_total}"
    ws.cell(59, 8).value = f"='{sheet}'!K{cur_traffic_total}"

    for i in range(16):
        r_v2 = 64 + i
        r_cur, _ = section_rows("推广数据", 43, 43, r_v2, i, 43 + i)
        ws.cell(r_v2, 3).value = f"='{sheet}'!Q{r_cur}"
        ws.cell(r_v2, 4).value = f"='{sheet}'!S{r_cur}"
        ws.cell(r_v2, 5).value = f"='{sheet}'!AA{r_cur}"
        ws.cell(r_v2, 6).value = f"='{sheet}'!AC{r_cur}"
    ws.cell(80, 3).value = f"='{sheet}'!Q{cur_promo_total}"
    ws.cell(80, 4).value = f"='{sheet}'!S{cur_promo_total}"
    ws.cell(80, 5).value = f"='{sheet}'!AA{cur_promo_total}"
    ws.cell(80, 6).value = f"='{sheet}'!AC{cur_promo_total}"

    for i in range(16):
        r_v2 = 85 + i
        r_cur, _ = section_rows("门店评分", 63, 63, r_v2, i, 63 + i)
        for c_v2, c_cur in zip([3, 4, 5, 6, 7, 8], ["N", "P", "Q", "X", "Z", "AA"]):
            ws.cell(r_v2, c_v2).value = f"='{sheet}'!{c_cur}{r_cur}"


def write_distance_and_paid(wb, prev_wb, distance, paid, stores: list[Store]) -> None:
    ws = wb["订单距离及实付区间"]
    prev = prev_wb["订单距离及实付区间"]
    pct_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 15, 16, 17, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29, 30, 31, 32, 33]
    prev_distance_maps = sheet_store_row_maps(prev, 5, 20, 2, 1)
    prev_paid_maps = sheet_store_row_maps(prev, 24, 39, 2, 1)
    for r in range(5, 21):
        code = store_code_by_sheet_row(ws, r, stores, 2, 1)
        prev_r = matched_row_by_code_name(code, ws.cell(r, 1).value, *prev_distance_maps)
        p = distance.get(code, {"total": {}, "mt": {}, "ele": {}})
        # Previous current values into prior columns.
        for cur_col, prev_col in [(3, 4), (6, 7), (9, 10), (14, 15), (17, 18), (20, 21), (25, 26), (28, 29), (31, 32)]:
            write(ws.cell(r, prev_col), prev.cell(prev_r, cur_col).value if prev_r else 0)
        vals = [
            (3, p["total"].get("0_1", 0.0)),
            (6, p["total"].get("1_3", 0.0)),
            (9, p["total"].get("gt3", 0.0)),
            (14, p["mt"].get("0_1", 0.0)),
            (17, p["mt"].get("1_3", 0.0)),
            (20, p["mt"].get("gt3", 0.0)),
            (25, p["ele"].get("0_1", 0.0)),
            (28, p["ele"].get("1_3", 0.0)),
            (31, p["ele"].get("gt3", 0.0)),
        ]
        for col, val in vals:
            write(ws.cell(r, col), val)
        for current_col, prior_col, delta_col in [(3, 4, 5), (6, 7, 8), (9, 10, 11), (14, 15, 16), (17, 18, 19), (20, 21, 22), (25, 26, 27), (28, 29, 30), (31, 32, 33)]:
            write(ws.cell(r, delta_col), diff(ws.cell(r, current_col).value, ws.cell(r, prior_col).value))

        r2 = r + 19
        paid_code = store_code_by_sheet_row(ws, r2, stores, 2, 1) or code
        prev_r2 = matched_row_by_code_name(paid_code, ws.cell(r2, 1).value, *prev_paid_maps)
        q = paid.get(paid_code, {"total": {}, "mt": {}, "ele": {}})
        for cur_col, prev_col in [(3, 4), (6, 7), (9, 10), (14, 15), (17, 18), (20, 21), (25, 26), (28, 29), (31, 32)]:
            write(ws.cell(r2, prev_col), prev.cell(prev_r2, cur_col).value if prev_r2 else 0)
        vals2 = [
            (3, q["total"].get("0_20", 0.0)),
            (6, q["total"].get("20_30", 0.0)),
            (9, q["total"].get("gt30", 0.0)),
            (14, q["mt"].get("0_20", 0.0)),
            (17, q["mt"].get("20_30", 0.0)),
            (20, q["mt"].get("gt30", 0.0)),
            (25, q["ele"].get("0_20", 0.0)),
            (28, q["ele"].get("20_30", 0.0)),
            (31, q["ele"].get("gt30", 0.0)),
        ]
        for col, val in vals2:
            write(ws.cell(r2, col), val)
        for current_col, prior_col, delta_col in [(3, 4, 5), (6, 7, 8), (9, 10, 11), (14, 15, 16), (17, 18, 19), (20, 21, 22), (25, 26, 27), (28, 29, 30), (31, 32, 33)]:
            write(ws.cell(r2, delta_col), diff(ws.cell(r2, current_col).value, ws.cell(r2, prior_col).value))
    for row in list(range(5, 21)) + list(range(24, 40)):
        for col in pct_cols:
            ws.cell(row, col).number_format = "0.0%"


def write_products(wb, prev_wb, single_rows, pkg_rows, prev_single_qty, prev_pkg_qty, total_store_days, new_packages) -> None:
    def set_num_formats(sheet, row_start: int, row_end: int, int_cols: list[int], pct_cols: list[int]) -> None:
        for row in range(row_start, row_end + 1):
            for col in int_cols:
                sheet.cell(row, col).number_format = "#,##0"
            for col in pct_cols:
                sheet.cell(row, col).number_format = "0.0%"

    ws = wb["商品销售排行-单品"]
    prev = prev_wb["商品销售排行-单品"]
    ws["G1"] = "上周销量"
    ws["H1"] = "环比"
    template_total_row = find_label_row(ws, "总计", None, 2) or 63
    prev_total_row = find_label_row(prev, "总计", None, 2) or 63
    existing_capacity = max(template_total_row - 2, 0)
    previous_capacity = max(prev_total_row - 2, 0)
    needed_capacity = max(len(single_rows), existing_capacity, previous_capacity)
    if needed_capacity > existing_capacity:
        rows_to_add = needed_capacity - existing_capacity
        ws.insert_rows(template_total_row, rows_to_add)
        for row in range(template_total_row, template_total_row + rows_to_add):
            copy_row_style(ws, template_total_row - 1, row, 13)
    total_row = 2 + needed_capacity
    clear_end = max(total_row - 1, template_total_row - 1)
    for row in range(2, clear_end + 1):
        for col in range(1, 9):
            ws.cell(row, col).value = None
        ws.cell(row, 12).value = None
        ws.cell(row, 13).value = None
    # Previous current product list as auxiliary.
    for idx, row in enumerate(range(2, prev_total_row)):
        ws.cell(row, 12).value = prev.cell(row, 3).value
        ws.cell(row, 13).value = prev.cell(row, 4).value
    for idx, item in enumerate(single_rows, start=2):
        prev_qty = prev_single_qty.get(item["name"], 0.0)
        values = [idx - 1, item["category"], item["name"], item["qty"], item["usd"], item["sales"], prev_qty, diff(item["qty"], prev_qty)]
        for col, value in enumerate(values, start=1):
            write(ws.cell(idx, col), value)
    qty_sum = sum(item["qty"] for item in single_rows)
    sales_sum = sum(item["sales"] for item in single_rows)
    prev_sum = scalar_num(prev.cell(prev_total_row, 4).value)
    write(ws.cell(total_row, 1), "总计")
    for col, value in {4: qty_sum, 5: safe_div(qty_sum, total_store_days or PERIOD_DAYS), 6: sales_sum, 7: prev_sum, 8: diff(qty_sum, prev_sum), 13: prev_sum}.items():
        write(ws.cell(total_row, col), value)
    set_num_formats(ws, 2, total_row, [4, 5, 6, 7, 8, 13], [])
    ws.column_dimensions["L"].hidden = True
    ws.column_dimensions["M"].hidden = True

    ws2 = wb["商品销售排行-套餐"]
    ws2["F1"] = "上周销量"
    ws2["I1"] = "上周销量占比"
    needed_rows = len(pkg_rows)
    total_row = 2 + needed_rows
    existing_total = 9
    if needed_rows > 7:
        ws2.insert_rows(existing_total, needed_rows - 7)
        for r in range(existing_total, total_row):
            copy_row_style(ws2, existing_total - 1, r, 13)
    ws2.cell(1, 12).value = "产品名"
    ws2.cell(1, 13).value = "销售量"
    for r in range(2, total_row):
        for c in range(1, 14):
            ws2.cell(r, c).value = None
    for i, (name, qty) in enumerate(prev_pkg_qty.items(), start=2):
        ws2.cell(i, 12).value = name
        ws2.cell(i, 13).value = qty
    total_qty = sum(item["qty"] for item in pkg_rows)
    total_sales = sum(item["sales"] for item in pkg_rows)
    prev_total = sum(prev_pkg_qty.get(item["name"], 0.0) for item in pkg_rows)
    for idx, item in enumerate(pkg_rows, start=2):
        prev_qty = prev_pkg_qty.get(item["name"], 0.0)
        share = safe_div(item["qty"], total_qty)
        prev_share = safe_div(prev_qty, prev_total)
        values = [
            idx - 1,
            item["name"],
            item["qty"],
            item["usd"],
            item["sales"],
            prev_qty,
            diff(item["qty"], prev_qty),
            share,
            prev_share,
            diff(share, prev_share),
        ]
        for col, value in enumerate(values, start=1):
            write(ws2.cell(idx, col), value)
    write(ws2.cell(total_row, 1), "总计")
    total_values = {3: total_qty, 4: safe_div(total_qty, total_store_days or PERIOD_DAYS), 5: total_sales, 6: prev_total, 7: diff(total_qty, prev_total), 8: 1, 9: 1, 10: 0}
    for col, value in total_values.items():
        write(ws2.cell(total_row, col), value)
    set_num_formats(ws2, 2, total_row, [3, 4, 5, 6, 7, 13], [8, 9, 10])
    ws2.column_dimensions["E"].width = max(float(ws2.column_dimensions["E"].width or 0), 12)
    ws2.column_dimensions["L"].hidden = True
    ws2.column_dimensions["M"].hidden = True


def write_complaints(wb, prev_wb, counts, top_rows) -> None:
    ws = wb["用户体验-客诉"]
    prev = prev_wb["用户体验-客诉"]
    cats = ["少餐具", "少菜品", "少酱料", "送错餐", "未看备注", "菜品口味", "菜品质量", "性价比低", "菜品异物", "客户原因", "物流问题", "服务态度差", "包装问题", "其他"]
    row_map = {ws.cell(r, 2).value: r for r in range(3, 17)}
    for cat in cats:
        r = row_map[cat]
        prev_val = prev.cell(r, 4).value
        cur_val = counts.get(cat, 0.0)
        write(ws.cell(r, 3), prev_val)
        write(ws.cell(r, 4), cur_val)
        write(ws.cell(r, 6), safe_div(cur_val, sum(counts.values())))
        write(ws.cell(r, 7), diff(cur_val, prev_val))
        ws.cell(r, 6).number_format = "0.0%"
    write(ws.cell(17, 3), sum((prev.cell(row_map[cat], 4).value or 0) for cat in cats))
    write(ws.cell(17, 4), sum(counts.values()))
    write(ws.cell(17, 7), diff(ws.cell(17, 4).value, ws.cell(17, 3).value))

    start = 23
    for r in range(start, 33):
        for c in range(1, 17):
            ws.cell(r, c).value = None
    for idx, item in enumerate(top_rows[:10], start=start):
        values = [item["store"], item["total"]] + [item.get(cat, 0.0) for cat in cats]
        for col, value in enumerate(values, start=1):
            write(ws.cell(idx, col), value)


def write_delivery(wb, prev_wb, mt_delivery, ele_delivery) -> None:
    ws = wb["用户体验-配送"]
    prev = prev_wb["用户体验-配送"]
    prev_maps = delivery_row_maps(prev)
    for row in range(3, 19):
        mt_id = norm_id(ws.cell(row, 4).value)
        ele_id = norm_id(ws.cell(row, 3).value)
        prev_row = matched_delivery_row(ws, row, prev_maps) or row
        # Map by IDs using ORA table indirectly from delivery rows.
        mt_value = None
        ele_value = None
        # Delivery dicts are keyed by store code, so lookup through row IDs.
        # The caller writes helper maps into hidden attributes via closures? Simpler: match below in main.
        write(ws.cell(row, 9), prev.cell(prev_row, 8).value)
        write(ws.cell(row, 16), prev.cell(prev_row, 15).value)
        # Current values are filled by main using ID-to-code maps after this function starts.


def fill_delivery_values(wb, prev_wb, mt_delivery, ele_delivery, mt_to_code, ele_to_code) -> None:
    ws = wb["用户体验-配送"]
    prev = prev_wb["用户体验-配送"]
    prev_maps = delivery_row_maps(prev)
    for row in range(3, 19):
        mt_code = mt_to_code.get(norm_id(ws.cell(row, 4).value))
        ele_code = ele_to_code.get(norm_id(ws.cell(row, 3).value))
        prev_row = matched_delivery_row(ws, row, prev_maps) or row
        mt_cur = mt_delivery.get(mt_code) if mt_code else None
        ele_cur = ele_delivery.get(ele_code) if ele_code else None
        write(ws.cell(row, 8), mt_cur)
        write(ws.cell(row, 9), prev.cell(prev_row, 8).value)
        write(ws.cell(row, 10), diff(mt_cur, ws.cell(row, 9).value))
        write(ws.cell(row, 15), ele_cur)
        write(ws.cell(row, 16), prev.cell(prev_row, 15).value)
        write(ws.cell(row, 17), diff(ele_cur, ws.cell(row, 16).value))


def copy_previous_week_to_previous_sheet(wb, prev_wb, prev_wb_format) -> None:
    ws = wb["上期"]
    prev_values = previous_period_sheet(prev_wb)
    prev_format = previous_period_sheet(prev_wb_format)
    layout = sheet_layout_from_xlsx(PREVIOUS, prev_format.title)

    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    source_max_row = max(prev_values.max_row, prev_format.max_row)
    source_max_col = max(prev_values.max_column, prev_format.max_column)
    max_row = max(ws.max_row, source_max_row)
    max_col = max(ws.max_column, source_max_col)
    blank_style = copy(ws.cell(1, 1)._style)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell._style = copy(blank_style)

    def attr_bool(attrs: dict[str, str], key: str) -> bool:
        return str(attrs.get(key, "")).lower() in {"1", "true"}

    for col_attrs in layout["cols"]:
        min_col = int(float(col_attrs.get("min", "1")))
        max_col_attr = int(float(col_attrs.get("max", str(min_col))))
        for idx in range(min_col, max_col_attr + 1):
            target = ws.column_dimensions[get_column_letter(idx)]
            if "width" in col_attrs:
                target.width = float(col_attrs["width"])
            target.hidden = attr_bool(col_attrs, "hidden")
            if "outlineLevel" in col_attrs:
                target.outlineLevel = int(float(col_attrs["outlineLevel"]))
            target.collapsed = attr_bool(col_attrs, "collapsed")
            target.bestFit = attr_bool(col_attrs, "bestFit")

    for idx, row_attrs in layout["rows"].items():
        target = ws.row_dimensions[idx]
        if "ht" in row_attrs:
            target.height = float(row_attrs["ht"])
        target.hidden = attr_bool(row_attrs, "hidden")
        if "outlineLevel" in row_attrs:
            target.outlineLevel = int(float(row_attrs["outlineLevel"]))
        target.collapsed = attr_bool(row_attrs, "collapsed")

    value_rows = prev_values.iter_rows(min_row=1, max_row=source_max_row, min_col=1, max_col=source_max_col)
    format_rows = prev_format.iter_rows(min_row=1, max_row=source_max_row, min_col=1, max_col=source_max_col)
    for r, (value_row, format_row) in enumerate(zip(value_rows, format_rows), start=1):
        for c, (value_cell, src_fmt) in enumerate(zip(value_row, format_row), start=1):
            dst = ws.cell(r, c)
            dst.value = clean_error_value(value_cell.value)
            if getattr(src_fmt, "has_style", False):
                dst.font = copy(src_fmt.font)
                dst.fill = copy(src_fmt.fill)
                dst.border = copy(src_fmt.border)
                dst.alignment = copy(src_fmt.alignment)
                dst.protection = copy(src_fmt.protection)
                dst.number_format = src_fmt.number_format

    for merged_range in layout["merges"]:
        ws.merge_cells(str(merged_range))

    ws.freeze_panes = layout["freeze"]


def validate_output(path: Path, stores: list[Store], metrics, totals, new_packages, paid_audit) -> dict[str, Any]:
    validation: dict[str, Any] = {}
    with zipfile.ZipFile(path) as zf:
        bad = zf.testzip()
        sheets = [name for name in zf.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]
        validation["zip_test"] = "ok" if bad is None else f"bad:{bad}"
        validation["worksheet_xml_count"] = len(sheets)
        validation["worksheet_xml_unique"] = len(sheets) == len(set(sheets))

    wb_formula = load_workbook(path, data_only=False)
    wb_values = load_workbook(path, data_only=True)
    validation["visible_sheets"] = [ws.title for ws in wb_formula.worksheets if ws.sheet_state == "visible"]
    v2_refs = []
    for row in range(3, 20):
        value = wb_formula["V2"].cell(row, 3).value
        if isinstance(value, str) and CURRENT_SHEET in value:
            v2_refs.append(row)
    validation["v2_refs_current_sheet_rows"] = v2_refs
    validation["previous_sheet_has_formulas"] = any(
        isinstance(wb_formula["上期"].cell(r, c).value, str) and wb_formula["上期"].cell(r, c).value.startswith("=")
        for r in range(1, 79)
        for c in range(1, 51)
    )
    validation["store_presence"] = {}
    checks = {
        CURRENT_SHEET: [(3, 18, 2), (23, 38, 2), (43, 58, 2), (63, 78, 2)],
        "V2": [(3, 18, 1), (23, 38, 1), (43, 58, 1), (64, 79, 1), (85, 100, 1)],
        "订单距离及实付区间": [(5, 20, 2), (24, 39, 2)],
    }
    for sheet, ranges in checks.items():
        ws = wb_formula[sheet]
        codes = set()
        for min_r, max_r, col in ranges:
            for r in range(min_r, max_r + 1):
                value = ws.cell(r, col).value
                if value:
                    codes.add(str(value))
        validation["store_presence"][sheet] = sorted(set(s.code for s in stores) - codes)
    validation["key_cells"] = {
        "current_total_sales": wb_values[CURRENT_SHEET]["E19"].value,
        "current_total_orders": wb_values[CURRENT_SHEET]["J19"].value,
        "current_mt_ad_spend": wb_values[CURRENT_SHEET]["Q59"].value,
        "current_ele_ad_spend": wb_values[CURRENT_SHEET]["AA59"].value,
        "single_top_product": wb_values["商品销售排行-单品"]["C2"].value,
        "package_total_qty": wb_values["商品销售排行-套餐"]["C" + str(2 + 7 + len(new_packages))].value,
        "complaint_total": wb_values["用户体验-客诉"]["D17"].value,
        "delivery_mt_first": wb_values["用户体验-配送"]["H3"].value,
        "delivery_ele_first": wb_values["用户体验-配送"]["O3"].value,
    }
    validation["new_packages"] = new_packages
    validation["paid_audit"] = paid_audit
    wb_formula.close()
    wb_values.close()
    return validation


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stores, mt_to_code, ele_to_code = build_store_list()
    metrics, totals = compute_metrics(stores, mt_to_code, ele_to_code)
    distance = compute_distance(mt_to_code, ele_to_code)
    paid, paid_audit = compute_paid_intervals(mt_to_code, ele_to_code, stores)

    prev_wb = load_workbook(PREVIOUS, data_only=True, read_only=True)
    prev_wb_format = load_workbook(PREVIOUS, data_only=False, read_only=True)
    single_rows, pkg_rows, prev_single_qty, prev_pkg_qty, new_packages = compute_products(prev_wb, totals["biz_days"] or PERIOD_DAYS)
    complaints, complaint_top = compute_complaints()
    mt_delivery, ele_delivery = compute_delivery(mt_to_code, ele_to_code)

    wb = load_workbook(TEMPLATE)
    ensure_current_sheet(wb)
    copy_previous_week_to_previous_sheet(wb, prev_wb, prev_wb_format)
    write_main_sheet(wb, prev_wb, stores, metrics, totals)
    write_v2(wb)
    write_distance_and_paid(wb, prev_wb, distance, paid, stores)
    write_products(wb, prev_wb, single_rows, pkg_rows, prev_single_qty, prev_pkg_qty, totals["biz_days"] or PERIOD_DAYS, new_packages)
    write_complaints(wb, prev_wb, complaints, complaint_top)
    fill_delivery_values(wb, prev_wb, mt_delivery, ele_delivery, mt_to_code, ele_to_code)
    write_business_analysis(wb)
    write_email_content_sheet(wb)
    refresh_period_labels(wb)
    apply_total_row_bold(wb)

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    output = OUT_DIR / os.environ.get("ORA_OUTPUT_NAME", f"ORA外送周报_{export_label(START, END)}.xlsx")
    wb.save(output)
    wb.close()
    prev_wb.close()
    prev_wb_format.close()

    expected_codes = set(s.code for s in stores)
    audit = {
        "output": str(output),
        "period": f"{START.date()}~{END.date()}",
        "stores": [{"code": s.code, "name": s.name, "mt_id": s.mt_id, "ele_id": s.ele_id} for s in stores],
        "new_stores": [],
        "missing_source_data": {
            "美团推广": sorted(expected_codes - totals["mt_promo_codes"]),
            "饿了么推广": sorted(expected_codes - totals["ele_promo_codes"]),
            "美团订单数据_匹配美团ID": "no matched rows" if paid_audit["mt_order_matched_rows"] == 0 else paid_audit["mt_order_matched_rows"],
            "饿了么订单数据_匹配饿了么ID": paid_audit["ele_order_matched_rows"],
        },
        "business_store_days": {code: metrics[code]["biz_days"] for code in expected_codes},
        "total_business_store_days": totals["biz_days"],
        "paid_exp_source": totals["paid_exp_source"],
        "new_packages": new_packages,
    }
    audit["validation"] = validate_output(output, stores, metrics, totals, new_packages, paid_audit)
    audit_path = WORK / "ora_report_audit.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
